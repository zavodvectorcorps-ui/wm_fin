from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Query, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Literal
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import io
import csv
import re
import zipfile
import shutil
import base64
import asyncio
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
import gspread
from google.oauth2.service_account import Credentials as ServiceAccountCredentials

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ.get('JWT_SECRET', 'wmfinance_secret')
JWT_ALGORITHM = "HS256"

app = FastAPI(title="WM Finance API")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============== MODELS ==============

class UserCreate(BaseModel):
    email: str
    password: str
    name: str
    role: Literal["owner", "accountant", "manager", "superadmin"] = "owner"

class UserLogin(BaseModel):
    email: str  # Can be email or login for superadmin
    password: str

# Superadmin credentials
SUPERADMIN_LOGIN = "admin"
SUPERADMIN_PASSWORD = "220066mm"
SUPERADMIN_ID = "superadmin-wmfinance-001"

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: str
    name: str
    role: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class Account(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    type: Literal["checking", "cash", "card", "savings"]
    currency: Literal["PLN", "EUR", "USD"] = "PLN"
    bank: Optional[str] = None
    initial_balance: float = 0
    current_balance: float = 0
    is_active: bool = True
    user_id: str = ""

class AccountCreate(BaseModel):
    name: str
    type: Literal["checking", "cash", "card", "savings"]
    currency: Literal["PLN", "EUR", "USD"] = "PLN"
    bank: Optional[str] = None
    initial_balance: float = 0

class Category(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    type: Literal["income", "expense"]
    group: str
    default_direction: Optional[str] = None
    is_active: bool = True
    user_id: str = ""

class CategoryCreate(BaseModel):
    name: str
    type: Literal["income", "expense"]
    group: str
    default_direction: Optional[str] = None

class BusinessDirection(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    color: str
    description: Optional[str] = None
    is_active: bool = True
    user_id: str = ""

class DirectionCreate(BaseModel):
    name: str
    color: str
    description: Optional[str] = None

class Contractor(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    type: Literal["client", "supplier", "employee", "other"]
    group: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    comment: Optional[str] = None
    is_active: bool = True
    user_id: str = ""

class ContractorCreate(BaseModel):
    name: str
    type: Literal["client", "supplier", "employee", "other"]
    group: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    comment: Optional[str] = None

class Transaction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    date: str
    type: Literal["income", "expense", "transfer"]
    amount: float
    currency: Literal["PLN", "EUR", "USD"] = "PLN"
    category_id: Optional[str] = None
    category_name: Optional[str] = None
    direction_id: str
    direction_name: Optional[str] = None
    account_id: str
    account_name: Optional[str] = None
    to_account_id: Optional[str] = None
    contractor_id: Optional[str] = None
    contractor_name: Optional[str] = None
    project_id: Optional[str] = None
    source: Literal["manual", "import", "telegram_bot"] = "manual"
    description: Optional[str] = None
    status: Literal["fact", "plan"] = "fact"
    is_recurring: bool = False
    balance_after: float = 0
    user_id: str = ""

class TransactionCreate(BaseModel):
    date: str
    type: Literal["income", "expense", "transfer"]
    amount: float
    currency: Literal["PLN", "EUR", "USD"] = "PLN"
    category_id: Optional[str] = None
    direction_id: str
    account_id: str
    to_account_id: Optional[str] = None
    contractor_id: Optional[str] = None
    project_id: Optional[str] = None
    description: Optional[str] = None
    status: Literal["fact", "plan"] = "fact"
    is_recurring: bool = False

class PlannedPayment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    type: Literal["income", "expense"]
    amount: float
    currency: Literal["PLN", "EUR", "USD"] = "PLN"
    category_id: Optional[str] = None
    category_name: Optional[str] = None
    contractor_id: Optional[str] = None
    contractor_name: Optional[str] = None
    direction_id: str
    direction_name: Optional[str] = None
    account_id: str
    account_name: Optional[str] = None
    status: Literal["pending", "paid", "overdue", "postponed", "cancelled"] = "pending"
    recurrence: Literal["none", "weekly", "monthly", "quarterly"] = "none"
    comment: Optional[str] = None
    linked_transaction_id: Optional[str] = None
    user_id: str = ""

class PlannedPaymentCreate(BaseModel):
    date: str
    type: Literal["income", "expense"]
    amount: float
    currency: Literal["PLN", "EUR", "USD"] = "PLN"
    category_id: Optional[str] = None
    contractor_id: Optional[str] = None
    direction_id: str
    account_id: str
    recurrence: Literal["none", "weekly", "monthly", "quarterly"] = "none"
    comment: Optional[str] = None

class Project(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    direction_id: str
    direction_name: Optional[str] = None
    contractor_id: Optional[str] = None
    contractor_name: Optional[str] = None
    planned_amount: float = 0
    actual_amount: float = 0
    status: Literal["active", "completed", "cancelled"] = "active"
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    comment: Optional[str] = None
    user_id: str = ""

class ProjectCreate(BaseModel):
    name: str
    direction_id: str
    contractor_id: Optional[str] = None
    planned_amount: float = 0
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    comment: Optional[str] = None

class AutoRule(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    pattern: str
    category_id: Optional[str] = None
    direction_id: Optional[str] = None
    contractor_id: Optional[str] = None
    is_active: bool = True
    user_id: str = ""

class AutoRuleCreate(BaseModel):
    pattern: str
    category_id: Optional[str] = None
    direction_id: Optional[str] = None
    contractor_id: Optional[str] = None

class BotTransactionRequest(BaseModel):
    text: str
    user_token: str
    date: Optional[str] = None

# ============== DOCUMENT MODEL ==============

class Document(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    document_date: Optional[str] = None
    type: Literal["invoice", "bank_statement", "payment_order", "act", "contract", "receipt", "other"] = "other"
    file_name: str
    file_url: str
    file_size: int = 0
    mime_type: str = ""
    transaction_id: Optional[str] = None
    contractor_id: Optional[str] = None
    contractor_name: Optional[str] = None
    direction_id: Optional[str] = None
    direction_name: Optional[str] = None
    period: Optional[str] = None  # YYYY-MM format
    status: Literal["linked", "pending"] = "pending"
    source: Literal["manual", "email", "telegram_bot"] = "manual"
    description: Optional[str] = None
    user_id: str = ""

class DocumentCreate(BaseModel):
    document_date: Optional[str] = None
    type: Literal["invoice", "bank_statement", "payment_order", "act", "contract", "receipt", "other"] = "other"
    transaction_id: Optional[str] = None
    contractor_id: Optional[str] = None
    direction_id: Optional[str] = None
    period: Optional[str] = None
    description: Optional[str] = None

class Notification(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    type: Literal["overdue_payment", "low_balance", "uncategorized_import", "document_pending"] = "overdue_payment"
    title: str
    message: str
    is_read: bool = False
    link: Optional[str] = None
    user_id: str = ""

# ============== ADESK MIGRATION MODELS ==============

class AdeskMigrationDraft(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    adesk_id: str  # Original ID from Adesk
    date: str
    type: Literal["income", "expense", "transfer"]
    amount: float
    currency: str = "PLN"
    category_adesk: Optional[str] = None  # Original category from Adesk
    category_id: Optional[str] = None  # Mapped WM Finance category
    category_name: Optional[str] = None
    project_adesk: Optional[str] = None  # Original project from Adesk
    direction_id: Optional[str] = None  # Mapped WM Finance direction
    direction_name: Optional[str] = None
    contractor_adesk: Optional[str] = None
    contractor_id: Optional[str] = None
    contractor_name: Optional[str] = None
    account_adesk: Optional[str] = None
    account_id: Optional[str] = None
    account_name: Optional[str] = None
    description: Optional[str] = None
    status: Literal["ready", "needs_review", "error", "imported"] = "needs_review"
    error_reason: Optional[str] = None
    user_id: str = ""
    batch_id: str = ""  # Migration batch ID

class AdeskConnectionTest(BaseModel):
    api_token: str

class AdeskMigrationStart(BaseModel):
    api_token: str
    date_from: str
    date_to: str
    migrate_transactions: bool = True
    migrate_contractors: bool = True
    migrate_projects: bool = True
    migrate_accounts: bool = True
    migrate_planned: bool = False

class AdeskDraftUpdate(BaseModel):
    category_id: Optional[str] = None
    direction_id: Optional[str] = None
    contractor_id: Optional[str] = None
    account_id: Optional[str] = None
    description: Optional[str] = None

class AdeskBulkUpdate(BaseModel):
    draft_ids: List[str]
    category_id: Optional[str] = None
    direction_id: Optional[str] = None
    contractor_id: Optional[str] = None

# ============== INTEGRATION SETTINGS MODELS ==============

class IntegrationSettings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str = ""
    # Telegram settings
    telegram_bot_token: Optional[str] = None
    telegram_chat_id: Optional[str] = None
    telegram_auto_summary: bool = False
    telegram_summary_schedule: Literal["daily", "weekly", "monday", "friday"] = "weekly"
    telegram_summary_time: str = "09:00"  # HH:MM format
    # Adesk settings
    adesk_api_token: Optional[str] = None
    # Other integrations can be added here
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class TelegramSettingsUpdate(BaseModel):
    telegram_bot_token: Optional[str] = None
    telegram_chat_id: Optional[str] = None
    telegram_auto_summary: Optional[bool] = None
    telegram_summary_schedule: Optional[Literal["daily", "weekly", "monday", "friday"]] = None
    telegram_summary_time: Optional[str] = None

class TelegramTestMessage(BaseModel):
    bot_token: str
    chat_id: str

# ============== AUTH HELPERS ==============

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

def create_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "user_id": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(days=30)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ============== AUTH ROUTES ==============

@api_router.post("/auth/register")
async def register(data: UserCreate):
    """Registration is disabled. Use admin panel to create users."""
    raise HTTPException(status_code=403, detail="Регистрация отключена. Обратитесь к администратору.")

@api_router.post("/auth/login")
async def login(data: UserLogin):
    # Check for superadmin login
    if data.email == SUPERADMIN_LOGIN and data.password == SUPERADMIN_PASSWORD:
        # Ensure superadmin exists in DB
        superadmin = await db.users.find_one({"id": SUPERADMIN_ID}, {"_id": 0})
        if not superadmin:
            # Create superadmin user
            superadmin_data = {
                "id": SUPERADMIN_ID,
                "email": "admin@wmfinance.local",
                "name": "Super Admin",
                "role": "superadmin",
                "created_at": datetime.now(timezone.utc).isoformat(),
                "password_hash": hash_password(SUPERADMIN_PASSWORD)
            }
            await db.users.insert_one(superadmin_data)
            await seed_user_data(SUPERADMIN_ID)
            superadmin = superadmin_data
        
        token = create_token(SUPERADMIN_ID, "admin@wmfinance.local", "superadmin")
        return {"token": token, "user": {"id": SUPERADMIN_ID, "email": "admin@wmfinance.local", "name": "Super Admin", "role": "superadmin"}}
    
    # Regular email-based login
    user = await db.users.find_one({"email": data.email}, {"_id": 0})
    if not user or not verify_password(data.password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_token(user["id"], user["email"], user["role"])
    return {"token": token, "user": {"id": user["id"], "email": user["email"], "name": user["name"], "role": user["role"]}}

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    user = await db.users.find_one({"id": current_user["user_id"]}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user

# ============== SEED DATA ==============

async def seed_user_data(user_id: str):
    # Business Directions
    directions = [
        {"id": str(uuid.uuid4()), "name": "Теплицы", "color": "blue", "description": "Производство и продажа теплиц", "is_active": True, "user_id": user_id},
        {"id": str(uuid.uuid4()), "name": "Сауны", "color": "orange", "description": "Производство и продажа саун", "is_active": True, "user_id": user_id},
        {"id": str(uuid.uuid4()), "name": "Купели", "color": "green", "description": "Производство и продажа купелей", "is_active": True, "user_id": user_id},
        {"id": str(uuid.uuid4()), "name": "Общее", "color": "gray", "description": "Общие операции бизнеса", "is_active": True, "user_id": user_id},
    ]
    await db.directions.insert_many(directions)
    
    # Income Categories
    income_categories = [
        {"id": str(uuid.uuid4()), "name": "Приход от клиентов", "type": "income", "group": "Выручка", "is_active": True, "user_id": user_id},
        {"id": str(uuid.uuid4()), "name": "Предоплата от клиентов", "type": "income", "group": "Выручка", "is_active": True, "user_id": user_id},
        {"id": str(uuid.uuid4()), "name": "Доплата по заказу", "type": "income", "group": "Выручка", "is_active": True, "user_id": user_id},
        {"id": str(uuid.uuid4()), "name": "Возврат от поставщика", "type": "income", "group": "Прочие доходы", "is_active": True, "user_id": user_id},
        {"id": str(uuid.uuid4()), "name": "Прочий приход", "type": "income", "group": "Прочие доходы", "is_active": True, "user_id": user_id},
    ]
    await db.categories.insert_many(income_categories)
    
    # Expense Categories
    expense_categories = [
        {"id": str(uuid.uuid4()), "name": "Закупка материалов", "type": "expense", "group": "Себестоимость", "is_active": True, "user_id": user_id},
        {"id": str(uuid.uuid4()), "name": "Производство", "type": "expense", "group": "Себестоимость", "is_active": True, "user_id": user_id},
        {"id": str(uuid.uuid4()), "name": "Логистика/Доставка", "type": "expense", "group": "Себестоимость", "is_active": True, "user_id": user_id},
        {"id": str(uuid.uuid4()), "name": "Аренда", "type": "expense", "group": "Операционные расходы", "is_active": True, "user_id": user_id},
        {"id": str(uuid.uuid4()), "name": "Зарплата", "type": "expense", "group": "Операционные расходы", "is_active": True, "user_id": user_id},
        {"id": str(uuid.uuid4()), "name": "Реклама и маркетинг", "type": "expense", "group": "Операционные расходы", "is_active": True, "user_id": user_id},
        {"id": str(uuid.uuid4()), "name": "Связь и интернет", "type": "expense", "group": "Операционные расходы", "is_active": True, "user_id": user_id},
        {"id": str(uuid.uuid4()), "name": "Офисные расходы", "type": "expense", "group": "Операционные расходы", "is_active": True, "user_id": user_id},
        {"id": str(uuid.uuid4()), "name": "Инструменты и оборудование", "type": "expense", "group": "Операционные расходы", "is_active": True, "user_id": user_id},
        {"id": str(uuid.uuid4()), "name": "Налоги", "type": "expense", "group": "Налоги и сборы", "is_active": True, "user_id": user_id},
        {"id": str(uuid.uuid4()), "name": "Бухгалтерские услуги", "type": "expense", "group": "Налоги и сборы", "is_active": True, "user_id": user_id},
        {"id": str(uuid.uuid4()), "name": "Прочий расход", "type": "expense", "group": "Прочие расходы", "is_active": True, "user_id": user_id},
    ]
    await db.categories.insert_many(expense_categories)
    
    # Default Accounts
    accounts = [
        {"id": str(uuid.uuid4()), "name": "Cash PL", "type": "cash", "currency": "PLN", "bank": None, "initial_balance": 0, "current_balance": 0, "is_active": True, "user_id": user_id},
        {"id": str(uuid.uuid4()), "name": "mBank PLN", "type": "checking", "currency": "PLN", "bank": "mBank", "initial_balance": 0, "current_balance": 0, "is_active": True, "user_id": user_id},
        {"id": str(uuid.uuid4()), "name": "mBank EUR", "type": "checking", "currency": "EUR", "bank": "mBank", "initial_balance": 0, "current_balance": 0, "is_active": True, "user_id": user_id},
    ]
    await db.accounts.insert_many(accounts)

# ============== ADMIN USER MANAGEMENT ==============

class AdminUserCreate(BaseModel):
    email: str
    password: str
    name: str
    role: Literal["owner", "accountant", "manager"] = "owner"

class AdminUserUpdate(BaseModel):
    email: Optional[str] = None
    password: Optional[str] = None
    name: Optional[str] = None
    role: Optional[Literal["owner", "accountant", "manager"]] = None

async def require_superadmin(current_user: dict = Depends(get_current_user)):
    """Dependency to check if user is superadmin"""
    if current_user.get("role") != "superadmin":
        raise HTTPException(status_code=403, detail="Только для супер-администратора")
    return current_user

@api_router.get("/admin/users")
async def get_all_users(current_user: dict = Depends(require_superadmin)):
    """Get all users (superadmin only)"""
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return users

@api_router.post("/admin/users")
async def create_user(data: AdminUserCreate, current_user: dict = Depends(require_superadmin)):
    """Create a new user (superadmin only)"""
    existing = await db.users.find_one({"email": data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email уже используется")
    
    user = User(email=data.email, name=data.name, role=data.role)
    user_dict = user.model_dump()
    user_dict["password_hash"] = hash_password(data.password)
    
    await db.users.insert_one(user_dict)
    
    # Seed initial data for new user
    await seed_user_data(user.id)
    
    return {"id": user.id, "email": user.email, "name": user.name, "role": user.role, "created_at": user.created_at}

@api_router.put("/admin/users/{user_id}")
async def update_user(user_id: str, data: AdminUserUpdate, current_user: dict = Depends(require_superadmin)):
    """Update a user (superadmin only)"""
    # Prevent editing superadmin
    if user_id == SUPERADMIN_ID:
        raise HTTPException(status_code=403, detail="Нельзя редактировать супер-администратора")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    
    update_data = {}
    if data.email:
        # Check if email is taken by another user
        existing = await db.users.find_one({"email": data.email, "id": {"$ne": user_id}}, {"_id": 0})
        if existing:
            raise HTTPException(status_code=400, detail="Email уже используется")
        update_data["email"] = data.email
    if data.name:
        update_data["name"] = data.name
    if data.role:
        update_data["role"] = data.role
    if data.password:
        update_data["password_hash"] = hash_password(data.password)
    
    if update_data:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    updated_user = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return updated_user

@api_router.delete("/admin/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(require_superadmin)):
    """Delete a user (superadmin only)"""
    # Prevent deleting superadmin
    if user_id == SUPERADMIN_ID:
        raise HTTPException(status_code=403, detail="Нельзя удалить супер-администратора")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    
    # Delete user and all their data
    await db.users.delete_one({"id": user_id})
    
    # Optionally delete user's data (transactions, accounts, etc.)
    # await db.transactions.delete_many({"user_id": user_id})
    # await db.accounts.delete_many({"user_id": user_id})
    # etc.
    
    return {"status": "deleted"}

# ============== ACCOUNTS ROUTES ==============

@api_router.get("/accounts", response_model=List[Account])
async def get_accounts(current_user: dict = Depends(get_current_user)):
    accounts = await db.accounts.find({"user_id": current_user["user_id"], "is_active": True}, {"_id": 0}).to_list(100)
    return accounts

@api_router.post("/accounts", response_model=Account)
async def create_account(data: AccountCreate, current_user: dict = Depends(get_current_user)):
    account = Account(**data.model_dump(), user_id=current_user["user_id"], current_balance=data.initial_balance)
    await db.accounts.insert_one(account.model_dump())
    return account

@api_router.put("/accounts/{account_id}", response_model=Account)
async def update_account(account_id: str, data: AccountCreate, current_user: dict = Depends(get_current_user)):
    result = await db.accounts.update_one(
        {"id": account_id, "user_id": current_user["user_id"]},
        {"$set": data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Account not found")
    account = await db.accounts.find_one({"id": account_id}, {"_id": 0})
    return account

@api_router.delete("/accounts/{account_id}")
async def delete_account(account_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.accounts.update_one(
        {"id": account_id, "user_id": current_user["user_id"]},
        {"$set": {"is_active": False}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Account not found")
    return {"status": "deleted"}

# ============== CATEGORIES ROUTES ==============

@api_router.get("/categories", response_model=List[Category])
async def get_categories(
    type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"user_id": current_user["user_id"], "is_active": True}
    if type:
        query["type"] = type
    categories = await db.categories.find(query, {"_id": 0}).to_list(100)
    return categories

@api_router.post("/categories", response_model=Category)
async def create_category(data: CategoryCreate, current_user: dict = Depends(get_current_user)):
    category = Category(**data.model_dump(), user_id=current_user["user_id"])
    await db.categories.insert_one(category.model_dump())
    return category

@api_router.put("/categories/{category_id}", response_model=Category)
async def update_category(category_id: str, data: CategoryCreate, current_user: dict = Depends(get_current_user)):
    result = await db.categories.update_one(
        {"id": category_id, "user_id": current_user["user_id"]},
        {"$set": data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    category = await db.categories.find_one({"id": category_id}, {"_id": 0})
    return category

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.categories.update_one(
        {"id": category_id, "user_id": current_user["user_id"]},
        {"$set": {"is_active": False}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    return {"status": "deleted"}

# ============== DIRECTIONS ROUTES ==============

@api_router.get("/directions", response_model=List[BusinessDirection])
async def get_directions(current_user: dict = Depends(get_current_user)):
    directions = await db.directions.find({"user_id": current_user["user_id"], "is_active": True}, {"_id": 0}).to_list(100)
    return directions

@api_router.post("/directions", response_model=BusinessDirection)
async def create_direction(data: DirectionCreate, current_user: dict = Depends(get_current_user)):
    direction = BusinessDirection(**data.model_dump(), user_id=current_user["user_id"])
    await db.directions.insert_one(direction.model_dump())
    return direction

@api_router.put("/directions/{direction_id}", response_model=BusinessDirection)
async def update_direction(direction_id: str, data: DirectionCreate, current_user: dict = Depends(get_current_user)):
    result = await db.directions.update_one(
        {"id": direction_id, "user_id": current_user["user_id"]},
        {"$set": data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Direction not found")
    direction = await db.directions.find_one({"id": direction_id}, {"_id": 0})
    return direction

@api_router.delete("/directions/{direction_id}")
async def delete_direction(direction_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.directions.update_one(
        {"id": direction_id, "user_id": current_user["user_id"]},
        {"$set": {"is_active": False}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Direction not found")
    return {"status": "deleted"}

# ============== CONTRACTORS ROUTES ==============

@api_router.get("/contractors", response_model=List[Contractor])
async def get_contractors(
    type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"user_id": current_user["user_id"], "is_active": True}
    if type:
        query["type"] = type
    contractors = await db.contractors.find(query, {"_id": 0}).to_list(500)
    return contractors

@api_router.get("/contractors/{contractor_id}")
async def get_contractor(contractor_id: str, current_user: dict = Depends(get_current_user)):
    contractor = await db.contractors.find_one({"id": contractor_id, "user_id": current_user["user_id"]}, {"_id": 0})
    if not contractor:
        raise HTTPException(status_code=404, detail="Contractor not found")
    
    # Get contractor's transactions
    transactions = await db.transactions.find(
        {"contractor_id": contractor_id, "user_id": current_user["user_id"]},
        {"_id": 0}
    ).sort("date", -1).to_list(100)
    
    # Calculate totals
    total_income = sum(t["amount"] for t in transactions if t["type"] == "income")
    total_expense = sum(t["amount"] for t in transactions if t["type"] == "expense")
    
    contractor["transactions"] = transactions
    contractor["total_income"] = total_income
    contractor["total_expense"] = total_expense
    
    return contractor

@api_router.post("/contractors", response_model=Contractor)
async def create_contractor(data: ContractorCreate, current_user: dict = Depends(get_current_user)):
    contractor = Contractor(**data.model_dump(), user_id=current_user["user_id"])
    await db.contractors.insert_one(contractor.model_dump())
    return contractor

@api_router.put("/contractors/{contractor_id}", response_model=Contractor)
async def update_contractor(contractor_id: str, data: ContractorCreate, current_user: dict = Depends(get_current_user)):
    result = await db.contractors.update_one(
        {"id": contractor_id, "user_id": current_user["user_id"]},
        {"$set": data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Contractor not found")
    contractor = await db.contractors.find_one({"id": contractor_id}, {"_id": 0})
    return contractor

@api_router.delete("/contractors/{contractor_id}")
async def delete_contractor(contractor_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.contractors.update_one(
        {"id": contractor_id, "user_id": current_user["user_id"]},
        {"$set": {"is_active": False}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Contractor not found")
    return {"status": "deleted"}

# ============== TRANSACTIONS ROUTES ==============

async def update_account_balance(account_id: str, user_id: str):
    """Recalculate account balance based on all transactions"""
    account = await db.accounts.find_one({"id": account_id, "user_id": user_id}, {"_id": 0})
    if not account:
        return
    
    initial = account.get("initial_balance", 0)
    
    # Sum all income transactions
    income_cursor = db.transactions.find({
        "account_id": account_id,
        "user_id": user_id,
        "type": "income",
        "status": "fact"
    }, {"_id": 0, "amount": 1})
    income_total = sum([t["amount"] async for t in income_cursor])
    
    # Sum all expense transactions
    expense_cursor = db.transactions.find({
        "account_id": account_id,
        "user_id": user_id,
        "type": "expense",
        "status": "fact"
    }, {"_id": 0, "amount": 1})
    expense_total = sum([t["amount"] async for t in expense_cursor])
    
    # Sum transfers out
    transfer_out_cursor = db.transactions.find({
        "account_id": account_id,
        "user_id": user_id,
        "type": "transfer",
        "status": "fact"
    }, {"_id": 0, "amount": 1})
    transfer_out_total = sum([t["amount"] async for t in transfer_out_cursor])
    
    # Sum transfers in
    transfer_in_cursor = db.transactions.find({
        "to_account_id": account_id,
        "user_id": user_id,
        "type": "transfer",
        "status": "fact"
    }, {"_id": 0, "amount": 1})
    transfer_in_total = sum([t["amount"] async for t in transfer_in_cursor])
    
    new_balance = initial + income_total - expense_total - transfer_out_total + transfer_in_total
    
    await db.accounts.update_one(
        {"id": account_id, "user_id": user_id},
        {"$set": {"current_balance": new_balance}}
    )

@api_router.get("/transactions", response_model=List[Transaction])
async def get_transactions(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    type: Optional[str] = None,
    status: Optional[str] = None,
    account_id: Optional[str] = None,
    direction_id: Optional[str] = None,
    category_id: Optional[str] = None,
    contractor_id: Optional[str] = None,
    source: Optional[str] = None,
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"user_id": current_user["user_id"]}
    
    if date_from:
        query["date"] = {"$gte": date_from}
    if date_to:
        if "date" in query:
            query["date"]["$lte"] = date_to
        else:
            query["date"] = {"$lte": date_to}
    if type:
        query["type"] = type
    if status:
        query["status"] = status
    if account_id:
        query["account_id"] = account_id
    if direction_id:
        query["direction_id"] = direction_id
    if category_id:
        query["category_id"] = category_id
    if contractor_id:
        query["contractor_id"] = contractor_id
    if source:
        query["source"] = source
    if search:
        query["description"] = {"$regex": search, "$options": "i"}
    
    transactions = await db.transactions.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return transactions

@api_router.post("/transactions", response_model=Transaction)
async def create_transaction(data: TransactionCreate, current_user: dict = Depends(get_current_user)):
    # Get related entities names
    category_name = None
    if data.category_id:
        cat = await db.categories.find_one({"id": data.category_id}, {"_id": 0, "name": 1})
        category_name = cat["name"] if cat else None
    
    direction = await db.directions.find_one({"id": data.direction_id}, {"_id": 0, "name": 1})
    direction_name = direction["name"] if direction else None
    
    account = await db.accounts.find_one({"id": data.account_id}, {"_id": 0, "name": 1, "current_balance": 1})
    account_name = account["name"] if account else None
    
    contractor_name = None
    if data.contractor_id:
        contractor = await db.contractors.find_one({"id": data.contractor_id}, {"_id": 0, "name": 1})
        contractor_name = contractor["name"] if contractor else None
    
    transaction = Transaction(
        **data.model_dump(),
        user_id=current_user["user_id"],
        category_name=category_name,
        direction_name=direction_name,
        account_name=account_name,
        contractor_name=contractor_name,
        source="manual"
    )
    
    await db.transactions.insert_one(transaction.model_dump())
    
    # Update account balance
    await update_account_balance(data.account_id, current_user["user_id"])
    if data.to_account_id:
        await update_account_balance(data.to_account_id, current_user["user_id"])
    
    # Get updated balance
    updated_account = await db.accounts.find_one({"id": data.account_id}, {"_id": 0, "current_balance": 1})
    transaction.balance_after = updated_account["current_balance"] if updated_account else 0
    
    # Update transaction with balance
    await db.transactions.update_one(
        {"id": transaction.id},
        {"$set": {"balance_after": transaction.balance_after}}
    )
    
    return transaction

@api_router.put("/transactions/{transaction_id}", response_model=Transaction)
async def update_transaction(transaction_id: str, data: TransactionCreate, current_user: dict = Depends(get_current_user)):
    old_transaction = await db.transactions.find_one({"id": transaction_id, "user_id": current_user["user_id"]}, {"_id": 0})
    if not old_transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    # Get related entities names
    category_name = None
    if data.category_id:
        cat = await db.categories.find_one({"id": data.category_id}, {"_id": 0, "name": 1})
        category_name = cat["name"] if cat else None
    
    direction = await db.directions.find_one({"id": data.direction_id}, {"_id": 0, "name": 1})
    direction_name = direction["name"] if direction else None
    
    account = await db.accounts.find_one({"id": data.account_id}, {"_id": 0, "name": 1})
    account_name = account["name"] if account else None
    
    contractor_name = None
    if data.contractor_id:
        contractor = await db.contractors.find_one({"id": data.contractor_id}, {"_id": 0, "name": 1})
        contractor_name = contractor["name"] if contractor else None
    
    update_data = data.model_dump()
    update_data["category_name"] = category_name
    update_data["direction_name"] = direction_name
    update_data["account_name"] = account_name
    update_data["contractor_name"] = contractor_name
    
    await db.transactions.update_one(
        {"id": transaction_id, "user_id": current_user["user_id"]},
        {"$set": update_data}
    )
    
    # Update account balances
    await update_account_balance(data.account_id, current_user["user_id"])
    if old_transaction.get("account_id") != data.account_id:
        await update_account_balance(old_transaction["account_id"], current_user["user_id"])
    if data.to_account_id:
        await update_account_balance(data.to_account_id, current_user["user_id"])
    
    transaction = await db.transactions.find_one({"id": transaction_id}, {"_id": 0})
    return transaction

@api_router.delete("/transactions/{transaction_id}")
async def delete_transaction(transaction_id: str, current_user: dict = Depends(get_current_user)):
    transaction = await db.transactions.find_one({"id": transaction_id, "user_id": current_user["user_id"]}, {"_id": 0})
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    await db.transactions.delete_one({"id": transaction_id})
    
    # Update account balance
    await update_account_balance(transaction["account_id"], current_user["user_id"])
    if transaction.get("to_account_id"):
        await update_account_balance(transaction["to_account_id"], current_user["user_id"])
    
    return {"status": "deleted"}

# ============== PLANNED PAYMENTS ROUTES ==============

@api_router.get("/planned-payments", response_model=List[PlannedPayment])
async def get_planned_payments(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    type: Optional[str] = None,
    status: Optional[str] = None,
    direction_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"user_id": current_user["user_id"]}
    
    if date_from:
        query["date"] = {"$gte": date_from}
    if date_to:
        if "date" in query:
            query["date"]["$lte"] = date_to
        else:
            query["date"] = {"$lte": date_to}
    if type:
        query["type"] = type
    if status:
        query["status"] = status
    if direction_id:
        query["direction_id"] = direction_id
    
    payments = await db.planned_payments.find(query, {"_id": 0}).sort("date", 1).to_list(500)
    
    # Auto-update overdue status
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    for payment in payments:
        if payment["status"] == "pending" and payment["date"] < today:
            await db.planned_payments.update_one(
                {"id": payment["id"]},
                {"$set": {"status": "overdue"}}
            )
            payment["status"] = "overdue"
    
    return payments

@api_router.post("/planned-payments", response_model=PlannedPayment)
async def create_planned_payment(data: PlannedPaymentCreate, current_user: dict = Depends(get_current_user)):
    # Get related entities names
    category_name = None
    if data.category_id:
        cat = await db.categories.find_one({"id": data.category_id}, {"_id": 0, "name": 1})
        category_name = cat["name"] if cat else None
    
    contractor_name = None
    if data.contractor_id:
        contractor = await db.contractors.find_one({"id": data.contractor_id}, {"_id": 0, "name": 1})
        contractor_name = contractor["name"] if contractor else None
    
    direction = await db.directions.find_one({"id": data.direction_id}, {"_id": 0, "name": 1})
    direction_name = direction["name"] if direction else None
    
    account = await db.accounts.find_one({"id": data.account_id}, {"_id": 0, "name": 1})
    account_name = account["name"] if account else None
    
    payment = PlannedPayment(
        **data.model_dump(),
        user_id=current_user["user_id"],
        category_name=category_name,
        contractor_name=contractor_name,
        direction_name=direction_name,
        account_name=account_name
    )
    
    await db.planned_payments.insert_one(payment.model_dump())
    return payment

@api_router.put("/planned-payments/{payment_id}", response_model=PlannedPayment)
async def update_planned_payment(payment_id: str, data: PlannedPaymentCreate, current_user: dict = Depends(get_current_user)):
    result = await db.planned_payments.update_one(
        {"id": payment_id, "user_id": current_user["user_id"]},
        {"$set": data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    payment = await db.planned_payments.find_one({"id": payment_id}, {"_id": 0})
    return payment

@api_router.put("/planned-payments/{payment_id}/status")
async def update_payment_status(
    payment_id: str,
    status: str = Query(...),
    current_user: dict = Depends(get_current_user)
):
    result = await db.planned_payments.update_one(
        {"id": payment_id, "user_id": current_user["user_id"]},
        {"$set": {"status": status}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    return {"status": "updated"}

@api_router.delete("/planned-payments/{payment_id}")
async def delete_planned_payment(payment_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.planned_payments.delete_one({"id": payment_id, "user_id": current_user["user_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    return {"status": "deleted"}

# ============== PROJECTS ROUTES ==============

@api_router.get("/projects", response_model=List[Project])
async def get_projects(
    status: Optional[str] = None,
    direction_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"user_id": current_user["user_id"]}
    if status:
        query["status"] = status
    if direction_id:
        query["direction_id"] = direction_id
    
    projects = await db.projects.find(query, {"_id": 0}).to_list(500)
    return projects

@api_router.get("/projects/{project_id}")
async def get_project(project_id: str, current_user: dict = Depends(get_current_user)):
    project = await db.projects.find_one({"id": project_id, "user_id": current_user["user_id"]}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Get project's transactions
    transactions = await db.transactions.find(
        {"project_id": project_id, "user_id": current_user["user_id"]},
        {"_id": 0}
    ).sort("date", -1).to_list(100)
    
    # Calculate actual amount
    actual_income = sum(t["amount"] for t in transactions if t["type"] == "income")
    actual_expense = sum(t["amount"] for t in transactions if t["type"] == "expense")
    
    project["transactions"] = transactions
    project["actual_amount"] = actual_income - actual_expense
    project["total_income"] = actual_income
    project["total_expense"] = actual_expense
    
    return project

@api_router.post("/projects", response_model=Project)
async def create_project(data: ProjectCreate, current_user: dict = Depends(get_current_user)):
    # Get related entities names
    direction = await db.directions.find_one({"id": data.direction_id}, {"_id": 0, "name": 1})
    direction_name = direction["name"] if direction else None
    
    contractor_name = None
    if data.contractor_id:
        contractor = await db.contractors.find_one({"id": data.contractor_id}, {"_id": 0, "name": 1})
        contractor_name = contractor["name"] if contractor else None
    
    project = Project(
        **data.model_dump(),
        user_id=current_user["user_id"],
        direction_name=direction_name,
        contractor_name=contractor_name
    )
    
    await db.projects.insert_one(project.model_dump())
    return project

@api_router.put("/projects/{project_id}", response_model=Project)
async def update_project(project_id: str, data: ProjectCreate, current_user: dict = Depends(get_current_user)):
    result = await db.projects.update_one(
        {"id": project_id, "user_id": current_user["user_id"]},
        {"$set": data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    return project

@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.projects.delete_one({"id": project_id, "user_id": current_user["user_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    return {"status": "deleted"}

# ============== AUTO RULES ROUTES ==============

@api_router.get("/auto-rules", response_model=List[AutoRule])
async def get_auto_rules(current_user: dict = Depends(get_current_user)):
    rules = await db.auto_rules.find({"user_id": current_user["user_id"]}, {"_id": 0}).to_list(100)
    return rules

@api_router.post("/auto-rules", response_model=AutoRule)
async def create_auto_rule(data: AutoRuleCreate, current_user: dict = Depends(get_current_user)):
    rule = AutoRule(**data.model_dump(), user_id=current_user["user_id"])
    await db.auto_rules.insert_one(rule.model_dump())
    return rule

@api_router.put("/auto-rules/{rule_id}", response_model=AutoRule)
async def update_auto_rule(rule_id: str, data: AutoRuleCreate, current_user: dict = Depends(get_current_user)):
    result = await db.auto_rules.update_one(
        {"id": rule_id, "user_id": current_user["user_id"]},
        {"$set": data.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Rule not found")
    rule = await db.auto_rules.find_one({"id": rule_id}, {"_id": 0})
    return rule

@api_router.delete("/auto-rules/{rule_id}")
async def delete_auto_rule(rule_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.auto_rules.delete_one({"id": rule_id, "user_id": current_user["user_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Rule not found")
    return {"status": "deleted"}

# ============== ANALYTICS ROUTES ==============

@api_router.get("/analytics/summary")
async def get_analytics_summary(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    direction_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"user_id": current_user["user_id"], "status": "fact"}
    
    if date_from:
        query["date"] = {"$gte": date_from}
    if date_to:
        if "date" in query:
            query["date"]["$lte"] = date_to
        else:
            query["date"] = {"$lte": date_to}
    if direction_id:
        query["direction_id"] = direction_id
    
    transactions = await db.transactions.find(query, {"_id": 0}).to_list(10000)
    
    total_income = sum(t["amount"] for t in transactions if t["type"] == "income")
    total_expense = sum(t["amount"] for t in transactions if t["type"] == "expense")
    profit = total_income - total_expense
    
    # By direction
    by_direction = {}
    for t in transactions:
        dir_name = t.get("direction_name", "Общее")
        if dir_name not in by_direction:
            by_direction[dir_name] = {"income": 0, "expense": 0, "profit": 0}
        if t["type"] == "income":
            by_direction[dir_name]["income"] += t["amount"]
        elif t["type"] == "expense":
            by_direction[dir_name]["expense"] += t["amount"]
        by_direction[dir_name]["profit"] = by_direction[dir_name]["income"] - by_direction[dir_name]["expense"]
    
    # By category
    income_by_category = {}
    expense_by_category = {}
    for t in transactions:
        cat_name = t.get("category_name", "Без категории")
        if t["type"] == "income":
            income_by_category[cat_name] = income_by_category.get(cat_name, 0) + t["amount"]
        elif t["type"] == "expense":
            expense_by_category[cat_name] = expense_by_category.get(cat_name, 0) + t["amount"]
    
    # Account balances
    accounts = await db.accounts.find({"user_id": current_user["user_id"], "is_active": True}, {"_id": 0}).to_list(100)
    total_balance = sum(a.get("current_balance", 0) for a in accounts)
    
    # Upcoming payments
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    upcoming_payments = await db.planned_payments.find(
        {"user_id": current_user["user_id"], "status": {"$in": ["pending", "overdue"]}, "date": {"$gte": today}},
        {"_id": 0}
    ).sort("date", 1).limit(5).to_list(5)
    
    overdue_payments = await db.planned_payments.find(
        {"user_id": current_user["user_id"], "status": "overdue"},
        {"_id": 0}
    ).to_list(100)
    
    return {
        "total_income": total_income,
        "total_expense": total_expense,
        "profit": profit,
        "total_balance": total_balance,
        "by_direction": by_direction,
        "income_by_category": income_by_category,
        "expense_by_category": expense_by_category,
        "accounts": accounts,
        "upcoming_payments": upcoming_payments,
        "overdue_payments": overdue_payments
    }

@api_router.get("/analytics/daily-balance")
async def get_daily_balance(
    date_from: str,
    date_to: str,
    current_user: dict = Depends(get_current_user)
):
    """Get daily balance changes for chart"""
    transactions = await db.transactions.find(
        {
            "user_id": current_user["user_id"],
            "status": "fact",
            "date": {"$gte": date_from, "$lte": date_to}
        },
        {"_id": 0}
    ).sort("date", 1).to_list(10000)
    
    # Get initial balance (sum of all accounts)
    accounts = await db.accounts.find({"user_id": current_user["user_id"], "is_active": True}, {"_id": 0}).to_list(100)
    
    # Calculate transactions before date_from to get starting balance
    prev_transactions = await db.transactions.find(
        {
            "user_id": current_user["user_id"],
            "status": "fact",
            "date": {"$lt": date_from}
        },
        {"_id": 0}
    ).to_list(10000)
    
    initial_balance = sum(a.get("initial_balance", 0) for a in accounts)
    for t in prev_transactions:
        if t["type"] == "income":
            initial_balance += t["amount"]
        elif t["type"] == "expense":
            initial_balance -= t["amount"]
    
    # Group by date
    daily = {}
    running_balance = initial_balance
    
    from datetime import datetime as dt
    start = dt.strptime(date_from, "%Y-%m-%d")
    end = dt.strptime(date_to, "%Y-%m-%d")
    current = start
    
    while current <= end:
        date_str = current.strftime("%Y-%m-%d")
        daily[date_str] = {"date": date_str, "balance": running_balance, "income": 0, "expense": 0}
        current += timedelta(days=1)
    
    for t in transactions:
        date_str = t["date"]
        if date_str in daily:
            if t["type"] == "income":
                daily[date_str]["income"] += t["amount"]
                running_balance += t["amount"]
            elif t["type"] == "expense":
                daily[date_str]["expense"] += t["amount"]
                running_balance -= t["amount"]
            daily[date_str]["balance"] = running_balance
    
    # Recalculate running balance
    running_balance = initial_balance
    result = []
    for date_str in sorted(daily.keys()):
        running_balance += daily[date_str]["income"] - daily[date_str]["expense"]
        daily[date_str]["balance"] = running_balance
        result.append(daily[date_str])
    
    return result

@api_router.get("/analytics/monthly")
async def get_monthly_analytics(
    year: int,
    current_user: dict = Depends(get_current_user)
):
    """Get monthly income/expense for the year"""
    transactions = await db.transactions.find(
        {
            "user_id": current_user["user_id"],
            "status": "fact",
            "date": {"$gte": f"{year}-01-01", "$lte": f"{year}-12-31"}
        },
        {"_id": 0}
    ).to_list(10000)
    
    months = {}
    for i in range(1, 13):
        month_key = f"{year}-{str(i).zfill(2)}"
        months[month_key] = {"month": month_key, "income": 0, "expense": 0, "profit": 0}
    
    for t in transactions:
        month_key = t["date"][:7]
        if month_key in months:
            if t["type"] == "income":
                months[month_key]["income"] += t["amount"]
            elif t["type"] == "expense":
                months[month_key]["expense"] += t["amount"]
            months[month_key]["profit"] = months[month_key]["income"] - months[month_key]["expense"]
    
    return list(months.values())

# ============== IMPORT ROUTES ==============

@api_router.post("/import/preview")
async def preview_import(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Preview CSV/XLSX import"""
    content = await file.read()
    
    rows = []
    if file.filename.endswith('.csv'):
        text = content.decode('utf-8')
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    elif file.filename.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                rows.append(dict(zip(headers, row)))
    
    # Return first 100 rows and available columns
    columns = list(rows[0].keys()) if rows else []
    
    return {
        "columns": columns,
        "preview": rows[:100],
        "total_rows": len(rows)
    }

@api_router.post("/import/process")
async def process_import(
    file: UploadFile = File(...),
    date_column: str = Query(...),
    amount_column: str = Query(...),
    description_column: str = Query(...),
    type_column: Optional[str] = None,
    account_id: str = Query(...),
    direction_id: str = Query(...),
    current_user: dict = Depends(get_current_user)
):
    """Process and import transactions from CSV/XLSX"""
    content = await file.read()
    
    rows = []
    if file.filename.endswith('.csv'):
        text = content.decode('utf-8')
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    elif file.filename.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                rows.append(dict(zip(headers, row)))
    
    # Get auto rules
    rules = await db.auto_rules.find({"user_id": current_user["user_id"], "is_active": True}, {"_id": 0}).to_list(100)
    
    # Get account and direction names
    account = await db.accounts.find_one({"id": account_id}, {"_id": 0, "name": 1})
    direction = await db.directions.find_one({"id": direction_id}, {"_id": 0, "name": 1})
    
    imported = []
    duplicates = []
    
    for row in rows:
        try:
            date_val = str(row.get(date_column, ""))
            amount_val = row.get(amount_column, 0)
            description = str(row.get(description_column, ""))
            
            # Parse date
            if isinstance(date_val, str) and date_val:
                # Try different formats
                for fmt in ["%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d"]:
                    try:
                        parsed_date = datetime.strptime(date_val.strip(), fmt)
                        date_str = parsed_date.strftime("%Y-%m-%d")
                        break
                    except ValueError:
                        date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            else:
                date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            
            # Parse amount
            if isinstance(amount_val, str):
                amount_val = amount_val.replace(",", ".").replace(" ", "")
            amount = abs(float(amount_val))
            
            # Determine type
            trans_type = "expense"
            if type_column and row.get(type_column):
                type_val = str(row.get(type_column, "")).lower()
                if "income" in type_val or "приход" in type_val or "+" in type_val:
                    trans_type = "income"
            elif float(amount_val) > 0:
                trans_type = "income"
            
            # Check for duplicates
            existing = await db.transactions.find_one({
                "user_id": current_user["user_id"],
                "date": date_str,
                "amount": amount,
                "description": description
            }, {"_id": 0})
            
            if existing:
                duplicates.append({"date": date_str, "amount": amount, "description": description})
                continue
            
            # Apply auto rules
            category_id = None
            category_name = None
            matched_direction_id = direction_id
            matched_direction_name = direction["name"] if direction else None
            matched = False
            
            for rule in rules:
                if rule["pattern"].lower() in description.lower():
                    if rule.get("category_id"):
                        category_id = rule["category_id"]
                        cat = await db.categories.find_one({"id": category_id}, {"_id": 0, "name": 1})
                        category_name = cat["name"] if cat else None
                    if rule.get("direction_id"):
                        matched_direction_id = rule["direction_id"]
                        dir_doc = await db.directions.find_one({"id": matched_direction_id}, {"_id": 0, "name": 1})
                        matched_direction_name = dir_doc["name"] if dir_doc else None
                    matched = True
                    break
            
            transaction = Transaction(
                date=date_str,
                type=trans_type,
                amount=amount,
                currency="PLN",
                category_id=category_id,
                category_name=category_name,
                direction_id=matched_direction_id,
                direction_name=matched_direction_name,
                account_id=account_id,
                account_name=account["name"] if account else None,
                description=description,
                source="import",
                status="fact",
                user_id=current_user["user_id"]
            )
            
            await db.transactions.insert_one(transaction.model_dump())
            imported.append({
                "id": transaction.id,
                "date": date_str,
                "type": trans_type,
                "amount": amount,
                "description": description,
                "category_name": category_name,
                "direction_name": matched_direction_name,
                "matched": matched
            })
            
        except Exception as e:
            logger.error(f"Error importing row: {e}")
            continue
    
    # Update account balance
    await update_account_balance(account_id, current_user["user_id"])
    
    return {
        "imported_count": len(imported),
        "duplicate_count": len(duplicates),
        "imported": imported,
        "duplicates": duplicates
    }

# ============== TELEGRAM BOT API ==============

@api_router.post("/bot/transaction")
async def bot_create_transaction(data: BotTransactionRequest):
    """Create transaction from Telegram bot text"""
    # Verify token
    try:
        payload = jwt.decode(data.user_token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload["user_id"]
    except (jwt.ExpiredSignatureError, jwt.InvalidTokenError):
        raise HTTPException(status_code=401, detail="Invalid token")
    
    text = data.text.lower()
    
    # Parse direction
    direction_name = "Общее"
    if "теплиц" in text:
        direction_name = "Теплицы"
    elif "саун" in text:
        direction_name = "Сауны"
    elif "купел" in text:
        direction_name = "Купели"
    
    direction = await db.directions.find_one({"user_id": user_id, "name": direction_name}, {"_id": 0})
    if not direction:
        direction = await db.directions.find_one({"user_id": user_id, "name": "Общее"}, {"_id": 0})
    
    # Parse type
    trans_type = "expense"
    if any(word in text for word in ["приход", "получил", "оплатили", "поступление"]):
        trans_type = "income"
    
    # Parse amount
    numbers = re.findall(r'\d+(?:[.,]\d+)?', text)
    amount = float(numbers[0].replace(",", ".")) if numbers else 0
    
    # Get default account
    account = await db.accounts.find_one({"user_id": user_id, "is_active": True}, {"_id": 0})
    if not account:
        raise HTTPException(status_code=400, detail="No accounts found")
    
    # Clean description
    description = text
    for word in ["теплицы", "теплица", "сауна", "сауны", "купель", "купели", "расход", "приход", "плачу", "получил"]:
        description = description.replace(word, "")
    for num in numbers:
        description = description.replace(num, "")
    description = " ".join(description.split()).strip()
    
    date_str = data.date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    transaction = Transaction(
        date=date_str,
        type=trans_type,
        amount=amount,
        currency="PLN",
        direction_id=direction["id"] if direction else "",
        direction_name=direction_name,
        account_id=account["id"],
        account_name=account["name"],
        description=description or "Операция из Telegram",
        source="telegram_bot",
        status="fact",
        user_id=user_id
    )
    
    await db.transactions.insert_one(transaction.model_dump())
    await update_account_balance(account["id"], user_id)
    
    return {
        "status": "created",
        "transaction": {
            "id": transaction.id,
            "date": transaction.date,
            "type": transaction.type,
            "amount": transaction.amount,
            "direction": direction_name,
            "description": transaction.description
        }
    }

@api_router.get("/bot/report")
async def bot_get_report(
    period: str = Query("week"),
    direction: str = Query("all"),
    user_token: str = Query(...)
):
    """Get text report for Telegram bot"""
    try:
        payload = jwt.decode(user_token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload["user_id"]
    except (jwt.ExpiredSignatureError, jwt.InvalidTokenError):
        raise HTTPException(status_code=401, detail="Invalid token")
    
    now = datetime.now(timezone.utc)
    
    if period == "week":
        date_from = (now - timedelta(days=7)).strftime("%Y-%m-%d")
    elif period == "month":
        date_from = (now - timedelta(days=30)).strftime("%Y-%m-%d")
    else:
        date_from = now.strftime("%Y-%m-01")
    
    date_to = now.strftime("%Y-%m-%d")
    
    query = {
        "user_id": user_id,
        "status": "fact",
        "date": {"$gte": date_from, "$lte": date_to}
    }
    
    if direction != "all":
        query["direction_name"] = direction
    
    transactions = await db.transactions.find(query, {"_id": 0}).to_list(10000)
    
    total_income = sum(t["amount"] for t in transactions if t["type"] == "income")
    total_expense = sum(t["amount"] for t in transactions if t["type"] == "expense")
    profit = total_income - total_expense
    
    accounts = await db.accounts.find({"user_id": user_id, "is_active": True}, {"_id": 0}).to_list(100)
    total_balance = sum(a.get("current_balance", 0) for a in accounts)
    
    report = f"""📊 Отчёт за {period}
    
💰 Доходы: {total_income:,.2f} zł
💸 Расходы: {total_expense:,.2f} zł
📈 Прибыль: {profit:,.2f} zł

💳 Остаток на счетах: {total_balance:,.2f} zł"""
    
    return {"report": report}

# ============== AI ASSISTANT ROUTES ==============

@api_router.post("/ai/chat")
async def ai_chat(
    message: str = Query(...),
    current_user: dict = Depends(get_current_user)
):
    """AI Assistant chat endpoint"""
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    
    api_key = os.environ.get("EMERGENT_LLM_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="AI not configured")
    
    # Get financial context
    today = datetime.now(timezone.utc)
    current_month_start = today.strftime("%Y-%m-01")
    prev_month_start = (today.replace(day=1) - timedelta(days=1)).strftime("%Y-%m-01")
    prev_month_end = (today.replace(day=1) - timedelta(days=1)).strftime("%Y-%m-%d")
    
    # Current month data
    current_transactions = await db.transactions.find({
        "user_id": current_user["user_id"],
        "status": "fact",
        "date": {"$gte": current_month_start}
    }, {"_id": 0}).to_list(10000)
    
    current_income = sum(t["amount"] for t in current_transactions if t["type"] == "income")
    current_expense = sum(t["amount"] for t in current_transactions if t["type"] == "expense")
    
    # Previous month data
    prev_transactions = await db.transactions.find({
        "user_id": current_user["user_id"],
        "status": "fact",
        "date": {"$gte": prev_month_start, "$lte": prev_month_end}
    }, {"_id": 0}).to_list(10000)
    
    prev_income = sum(t["amount"] for t in prev_transactions if t["type"] == "income")
    prev_expense = sum(t["amount"] for t in prev_transactions if t["type"] == "expense")
    
    # By direction
    by_direction = {}
    for t in current_transactions:
        dir_name = t.get("direction_name", "Общее")
        if dir_name not in by_direction:
            by_direction[dir_name] = {"income": 0, "expense": 0}
        if t["type"] == "income":
            by_direction[dir_name]["income"] += t["amount"]
        elif t["type"] == "expense":
            by_direction[dir_name]["expense"] += t["amount"]
    
    # Top categories
    expense_by_cat = {}
    for t in current_transactions:
        if t["type"] == "expense":
            cat = t.get("category_name", "Без категории")
            expense_by_cat[cat] = expense_by_cat.get(cat, 0) + t["amount"]
    
    top_expenses = sorted(expense_by_cat.items(), key=lambda x: x[1], reverse=True)[:10]
    
    # Accounts
    accounts = await db.accounts.find({"user_id": current_user["user_id"], "is_active": True}, {"_id": 0}).to_list(100)
    
    # Planned payments
    upcoming = await db.planned_payments.find({
        "user_id": current_user["user_id"],
        "status": {"$in": ["pending", "overdue"]}
    }, {"_id": 0}).sort("date", 1).limit(10).to_list(10)
    
    overdue = [p for p in upcoming if p["status"] == "overdue"]
    
    context = f"""Контекст финансовых данных компании WM Finance (теплицы, сауны, купели):

Текущий месяц ({current_month_start[:7]}):
- Доходы: {current_income:,.2f} PLN
- Расходы: {current_expense:,.2f} PLN
- Прибыль: {current_income - current_expense:,.2f} PLN

Прошлый месяц:
- Доходы: {prev_income:,.2f} PLN
- Расходы: {prev_expense:,.2f} PLN
- Прибыль: {prev_income - prev_expense:,.2f} PLN

По направлениям бизнеса (текущий месяц):
{chr(10).join([f"- {k}: доходы {v['income']:,.2f}, расходы {v['expense']:,.2f}, прибыль {v['income']-v['expense']:,.2f}" for k, v in by_direction.items()])}

Топ расходов:
{chr(10).join([f"- {cat}: {amt:,.2f} PLN" for cat, amt in top_expenses])}

Счета:
{chr(10).join([f"- {a['name']}: {a['current_balance']:,.2f} {a['currency']}" for a in accounts])}

Ближайшие платежи:
{chr(10).join([f"- {p['date']}: {p['type']} {p['amount']:,.2f} PLN ({p['status']})" for p in upcoming[:5]])}

Просроченных платежей: {len(overdue)}
"""
    
    system_message = """Ты финансовый ИИ-ассистент компании WM Finance. Компания занимается производством и продажей теплиц, саун и купелей в Польше.

Отвечай на русском языке. Используй данные из контекста для ответов на вопросы о финансах.
Форматируй суммы с разделителем тысяч и 2 знаками после запятой.
Будь кратким и по существу.

Если пользователь просит добавить операцию, сформируй JSON с данными операции в формате:
{"action": "create_transaction", "data": {"type": "income/expense", "amount": число, "direction": "название", "description": "описание"}}
"""
    
    chat = LlmChat(
        api_key=api_key,
        session_id=f"wmfinance_{current_user['user_id']}_{today.strftime('%Y%m%d%H%M')}",
        system_message=system_message
    ).with_model("anthropic", "claude-sonnet-4-5-20250929")
    
    user_message = UserMessage(text=f"{context}\n\nВопрос пользователя: {message}")
    
    try:
        response = await chat.send_message(user_message)
        return {"response": response}
    except Exception as e:
        logger.error(f"AI error: {e}")
        raise HTTPException(status_code=500, detail="AI service error")

# ============== DOCUMENTS ROUTES ==============

# Create uploads directory
UPLOADS_DIR = ROOT_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)

@api_router.get("/documents", response_model=List[Document])
async def get_documents(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    type: Optional[str] = None,
    status: Optional[str] = None,
    direction_id: Optional[str] = None,
    period: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"user_id": current_user["user_id"]}
    
    if date_from:
        query["document_date"] = {"$gte": date_from}
    if date_to:
        if "document_date" in query:
            query["document_date"]["$lte"] = date_to
        else:
            query["document_date"] = {"$lte": date_to}
    if type:
        query["type"] = type
    if status:
        query["status"] = status
    if direction_id:
        query["direction_id"] = direction_id
    if period:
        query["period"] = period
    
    documents = await db.documents.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return documents

@api_router.get("/documents/pending", response_model=List[Document])
async def get_pending_documents(current_user: dict = Depends(get_current_user)):
    """Get documents that need processing (not linked to transactions)"""
    documents = await db.documents.find(
        {"user_id": current_user["user_id"], "status": "pending"},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return documents

@api_router.post("/documents/upload")
async def upload_document(
    file: UploadFile = File(...),
    document_date: Optional[str] = Form(None),
    type: str = Form("other"),
    direction_id: Optional[str] = Form(None),
    contractor_id: Optional[str] = Form(None),
    transaction_id: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    current_user: dict = Depends(get_current_user)
):
    """Upload a document file"""
    # Validate file type
    allowed_types = [".pdf", ".png", ".jpg", ".jpeg", ".xlsx", ".xls"]
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in allowed_types:
        raise HTTPException(status_code=400, detail=f"File type {file_ext} not allowed")
    
    # Generate unique filename
    file_id = str(uuid.uuid4())
    safe_filename = f"{file_id}{file_ext}"
    file_path = UPLOADS_DIR / safe_filename
    
    # Save file
    content = await file.read()
    file_size = len(content)
    
    with open(file_path, "wb") as f:
        f.write(content)
    
    # Get related entity names
    direction_name = None
    if direction_id:
        direction = await db.directions.find_one({"id": direction_id}, {"_id": 0, "name": 1})
        direction_name = direction["name"] if direction else None
    
    contractor_name = None
    if contractor_id:
        contractor = await db.contractors.find_one({"id": contractor_id}, {"_id": 0, "name": 1})
        contractor_name = contractor["name"] if contractor else None
    
    # Determine period from document_date
    period = None
    if document_date:
        period = document_date[:7]  # YYYY-MM
    
    # Determine status
    status = "linked" if transaction_id else "pending"
    
    document = Document(
        document_date=document_date,
        type=type,
        file_name=file.filename,
        file_url=f"/api/documents/file/{safe_filename}",
        file_size=file_size,
        mime_type=file.content_type or "",
        transaction_id=transaction_id,
        contractor_id=contractor_id,
        contractor_name=contractor_name,
        direction_id=direction_id,
        direction_name=direction_name,
        period=period,
        status=status,
        source="manual",
        description=description,
        user_id=current_user["user_id"]
    )
    
    await db.documents.insert_one(document.model_dump())
    
    return document

@api_router.get("/documents/file/{filename}")
async def get_document_file(filename: str):
    """Serve uploaded document file"""
    file_path = UPLOADS_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    # Determine content type
    ext = os.path.splitext(filename)[1].lower()
    content_types = {
        ".pdf": "application/pdf",
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xls": "application/vnd.ms-excel"
    }
    content_type = content_types.get(ext, "application/octet-stream")
    
    with open(file_path, "rb") as f:
        content = f.read()
    
    return StreamingResponse(
        io.BytesIO(content),
        media_type=content_type,
        headers={"Content-Disposition": f"inline; filename={filename}"}
    )

@api_router.put("/documents/{document_id}")
async def update_document(
    document_id: str,
    data: DocumentCreate,
    current_user: dict = Depends(get_current_user)
):
    """Update document metadata"""
    update_data = data.model_dump(exclude_unset=True)
    
    # Get related entity names
    if data.direction_id:
        direction = await db.directions.find_one({"id": data.direction_id}, {"_id": 0, "name": 1})
        update_data["direction_name"] = direction["name"] if direction else None
    
    if data.contractor_id:
        contractor = await db.contractors.find_one({"id": data.contractor_id}, {"_id": 0, "name": 1})
        update_data["contractor_name"] = contractor["name"] if contractor else None
    
    # Update status based on transaction_id
    update_data["status"] = "linked" if data.transaction_id else "pending"
    
    result = await db.documents.update_one(
        {"id": document_id, "user_id": current_user["user_id"]},
        {"$set": update_data}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Document not found")
    
    document = await db.documents.find_one({"id": document_id}, {"_id": 0})
    return document

@api_router.delete("/documents/{document_id}")
async def delete_document(document_id: str, current_user: dict = Depends(get_current_user)):
    """Delete a document"""
    document = await db.documents.find_one(
        {"id": document_id, "user_id": current_user["user_id"]},
        {"_id": 0}
    )
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Delete file
    if document.get("file_url"):
        filename = document["file_url"].split("/")[-1]
        file_path = UPLOADS_DIR / filename
        if file_path.exists():
            file_path.unlink()
    
    await db.documents.delete_one({"id": document_id})
    return {"status": "deleted"}

@api_router.post("/documents/{document_id}/link-transaction")
async def link_document_to_transaction(
    document_id: str,
    transaction_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Link a document to a transaction"""
    # Verify document exists
    document = await db.documents.find_one(
        {"id": document_id, "user_id": current_user["user_id"]},
        {"_id": 0}
    )
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Verify transaction exists
    transaction = await db.transactions.find_one(
        {"id": transaction_id, "user_id": current_user["user_id"]},
        {"_id": 0}
    )
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    # Update document
    await db.documents.update_one(
        {"id": document_id},
        {"$set": {
            "transaction_id": transaction_id,
            "status": "linked"
        }}
    )
    
    return {"status": "linked", "document_id": document_id, "transaction_id": transaction_id}

@api_router.delete("/documents/{document_id}/unlink")
async def unlink_document_from_transaction(
    document_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Unlink a document from its transaction"""
    await db.documents.update_one(
        {"id": document_id, "user_id": current_user["user_id"]},
        {"$set": {"transaction_id": None, "status": "pending"}}
    )
    return {"status": "unlinked"}

@api_router.get("/transactions/{transaction_id}/documents")
async def get_transaction_documents(
    transaction_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get all documents linked to a transaction"""
    documents = await db.documents.find(
        {"transaction_id": transaction_id, "user_id": current_user["user_id"]},
        {"_id": 0}
    ).to_list(100)
    return documents

@api_router.get("/documents/export")
async def export_documents(
    period: str = Query(..., description="Period in YYYY-MM format"),
    types: Optional[str] = Query(None, description="Comma-separated document types"),
    current_user: dict = Depends(get_current_user)
):
    """Export documents as ZIP archive"""
    query = {"user_id": current_user["user_id"], "period": period}
    
    if types:
        type_list = types.split(",")
        query["type"] = {"$in": type_list}
    
    documents = await db.documents.find(query, {"_id": 0}).to_list(1000)
    
    if not documents:
        raise HTTPException(status_code=404, detail="No documents found for export")
    
    # Create ZIP in memory
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for doc in documents:
            if doc.get("file_url"):
                filename = doc["file_url"].split("/")[-1]
                file_path = UPLOADS_DIR / filename
                
                if file_path.exists():
                    # Determine folder based on document type
                    folder = "прочее"
                    if doc["type"] in ["invoice", "receipt"]:
                        folder = "расходы" if doc.get("transaction_id") else "доходы"
                    elif doc["type"] == "bank_statement":
                        folder = "выписки"
                    elif doc["type"] in ["contract", "act"]:
                        folder = "договоры"
                    
                    # Add to ZIP with folder structure
                    archive_name = f"{folder}/{doc['file_name']}"
                    zip_file.write(file_path, archive_name)
    
    zip_buffer.seek(0)
    
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="documents_{period}.zip"'
        }
    )

@api_router.get("/documents/by-transaction/{transaction_id}", response_model=List[Document])
async def get_documents_by_transaction(
    transaction_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get documents linked to a specific transaction"""
    documents = await db.documents.find(
        {"user_id": current_user["user_id"], "transaction_id": transaction_id},
        {"_id": 0}
    ).to_list(100)
    return documents

# ============== NOTIFICATIONS ROUTES ==============

@api_router.get("/notifications")
async def get_notifications(current_user: dict = Depends(get_current_user)):
    """Get user notifications"""
    # Generate dynamic notifications
    notifications = []
    
    # Check for overdue payments
    overdue_payments = await db.planned_payments.find(
        {"user_id": current_user["user_id"], "status": "overdue"},
        {"_id": 0}
    ).to_list(100)
    
    if overdue_payments:
        notifications.append({
            "id": "overdue_payments",
            "type": "overdue_payment",
            "title": f"Просроченные платежи: {len(overdue_payments)}",
            "message": f"У вас {len(overdue_payments)} просроченных платежей на сумму {sum(p['amount'] for p in overdue_payments):,.2f} PLN",
            "is_read": False,
            "link": "/planning/calendar",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    # Check for pending documents
    pending_docs = await db.documents.count_documents(
        {"user_id": current_user["user_id"], "status": "pending"}
    )
    
    if pending_docs > 0:
        notifications.append({
            "id": "pending_docs",
            "type": "document_pending",
            "title": f"Документы без привязки: {pending_docs}",
            "message": f"{pending_docs} документов требуют обработки",
            "is_read": False,
            "link": "/documents?status=pending",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    # Check for low balance accounts
    accounts = await db.accounts.find(
        {"user_id": current_user["user_id"], "is_active": True},
        {"_id": 0}
    ).to_list(100)
    
    for account in accounts:
        if account.get("current_balance", 0) < 0:
            notifications.append({
                "id": f"low_balance_{account['id']}",
                "type": "low_balance",
                "title": f"Отрицательный баланс: {account['name']}",
                "message": f"Баланс счёта {account['name']}: {account['current_balance']:,.2f} {account['currency']}",
                "is_read": False,
                "link": "/settings",
                "created_at": datetime.now(timezone.utc).isoformat()
            })
    
    # Get stored notifications
    stored = await db.notifications.find(
        {"user_id": current_user["user_id"]},
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    
    notifications.extend(stored)
    
    return {
        "notifications": notifications,
        "unread_count": len([n for n in notifications if not n.get("is_read", False)])
    }

@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, current_user: dict = Depends(get_current_user)):
    """Mark notification as read"""
    await db.notifications.update_one(
        {"id": notification_id, "user_id": current_user["user_id"]},
        {"$set": {"is_read": True}}
    )
    return {"status": "ok"}

# ============== ANALYTICS P&L ==============

@api_router.get("/analytics/pnl")
async def get_pnl_report(
    date_from: str,
    date_to: str,
    direction_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get Profit & Loss report"""
    query = {
        "user_id": current_user["user_id"],
        "status": "fact",
        "date": {"$gte": date_from, "$lte": date_to}
    }
    
    if direction_id:
        query["direction_id"] = direction_id
    
    transactions = await db.transactions.find(query, {"_id": 0}).to_list(10000)
    categories = await db.categories.find({"user_id": current_user["user_id"]}, {"_id": 0}).to_list(100)
    
    # Group by category and calculate totals
    income_groups = {}
    expense_groups = {}
    
    for cat in categories:
        group = cat["group"]
        if cat["type"] == "income":
            if group not in income_groups:
                income_groups[group] = {"items": {}, "total": 0}
            income_groups[group]["items"][cat["name"]] = 0
        else:
            if group not in expense_groups:
                expense_groups[group] = {"items": {}, "total": 0}
            expense_groups[group]["items"][cat["name"]] = 0
    
    # Calculate amounts
    total_income = 0
    total_expense = 0
    
    for t in transactions:
        cat_name = t.get("category_name", "Без категории")
        cat = next((c for c in categories if c["name"] == cat_name), None)
        
        if t["type"] == "income":
            total_income += t["amount"]
            if cat:
                group = cat["group"]
                if group in income_groups:
                    income_groups[group]["items"][cat_name] = income_groups[group]["items"].get(cat_name, 0) + t["amount"]
                    income_groups[group]["total"] += t["amount"]
        elif t["type"] == "expense":
            total_expense += t["amount"]
            if cat:
                group = cat["group"]
                if group in expense_groups:
                    expense_groups[group]["items"][cat_name] = expense_groups[group]["items"].get(cat_name, 0) + t["amount"]
                    expense_groups[group]["total"] += t["amount"]
    
    return {
        "period": {"from": date_from, "to": date_to},
        "income": {
            "total": total_income,
            "groups": income_groups
        },
        "expense": {
            "total": total_expense,
            "groups": expense_groups
        },
        "gross_profit": total_income - total_expense,
        "net_profit": total_income - total_expense  # Simplified, same as gross for now
    }

@api_router.get("/analytics/cashflow")
async def get_cashflow_report(
    year: int,
    direction_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get Cash Flow report by month"""
    query = {
        "user_id": current_user["user_id"],
        "status": "fact",
        "date": {"$gte": f"{year}-01-01", "$lte": f"{year}-12-31"}
    }
    
    if direction_id:
        query["direction_id"] = direction_id
    
    transactions = await db.transactions.find(query, {"_id": 0}).to_list(10000)
    
    # Initialize months
    months = []
    for i in range(1, 13):
        months.append({
            "month": f"{year}-{str(i).zfill(2)}",
            "income": 0,
            "expense": 0,
            "net": 0,
            "by_category": {}
        })
    
    # Calculate by month
    for t in transactions:
        month_idx = int(t["date"][5:7]) - 1
        cat_name = t.get("category_name", "Без категории")
        
        if t["type"] == "income":
            months[month_idx]["income"] += t["amount"]
        elif t["type"] == "expense":
            months[month_idx]["expense"] += t["amount"]
        
        if cat_name not in months[month_idx]["by_category"]:
            months[month_idx]["by_category"][cat_name] = 0
        
        amount = t["amount"] if t["type"] == "income" else -t["amount"]
        months[month_idx]["by_category"][cat_name] += amount
        months[month_idx]["net"] = months[month_idx]["income"] - months[month_idx]["expense"]
    
    return {
        "year": year,
        "months": months,
        "total_income": sum(m["income"] for m in months),
        "total_expense": sum(m["expense"] for m in months),
        "net_cashflow": sum(m["net"] for m in months)
    }

# ============== ANALYTICS BALANCE ==============

@api_router.get("/analytics/balance")
async def get_balance_report(
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get Balance Sheet report - Assets and Liabilities"""
    if not date_to:
        date_to = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Get all accounts with balances
    accounts = await db.accounts.find(
        {"user_id": current_user["user_id"], "is_active": True},
        {"_id": 0}
    ).to_list(100)
    
    # Group accounts by type and currency
    assets = {
        "cash": [],
        "checking": [],
        "card": [],
        "savings": []
    }
    
    total_by_currency = {}
    
    for account in accounts:
        acc_type = account.get("type", "checking")
        currency = account.get("currency", "PLN")
        balance = account.get("current_balance", 0)
        
        assets[acc_type].append({
            "name": account["name"],
            "balance": balance,
            "currency": currency,
            "bank": account.get("bank")
        })
        
        if currency not in total_by_currency:
            total_by_currency[currency] = 0
        total_by_currency[currency] += balance
    
    # Calculate total in PLN (simplified - no conversion)
    total_assets = sum(a.get("current_balance", 0) for a in accounts)
    
    # Get pending planned payments (liabilities)
    pending_expenses = await db.planned_payments.find(
        {
            "user_id": current_user["user_id"],
            "type": "expense",
            "status": {"$in": ["pending", "overdue"]}
        },
        {"_id": 0}
    ).to_list(500)
    
    pending_income = await db.planned_payments.find(
        {
            "user_id": current_user["user_id"],
            "type": "income",
            "status": {"$in": ["pending", "overdue"]}
        },
        {"_id": 0}
    ).to_list(500)
    
    total_liabilities = sum(p["amount"] for p in pending_expenses)
    total_receivables = sum(p["amount"] for p in pending_income)
    
    return {
        "date": date_to,
        "assets": {
            "cash": assets["cash"],
            "checking": assets["checking"],
            "card": assets["card"],
            "savings": assets["savings"],
            "total": total_assets,
            "by_currency": total_by_currency
        },
        "liabilities": {
            "pending_payments": pending_expenses[:20],
            "total": total_liabilities
        },
        "receivables": {
            "pending_income": pending_income[:20],
            "total": total_receivables
        },
        "net_worth": total_assets - total_liabilities + total_receivables
    }

@api_router.get("/analytics/expense-analysis")
async def get_expense_analysis(
    date_from: str,
    date_to: str,
    direction_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Detailed expense analysis with trends"""
    query = {
        "user_id": current_user["user_id"],
        "status": "fact",
        "type": "expense",
        "date": {"$gte": date_from, "$lte": date_to}
    }
    
    if direction_id:
        query["direction_id"] = direction_id
    
    transactions = await db.transactions.find(query, {"_id": 0}).to_list(10000)
    
    # By category
    by_category = {}
    by_direction = {}
    by_contractor = {}
    daily_expenses = {}
    
    for t in transactions:
        cat_name = t.get("category_name", "Без категории")
        dir_name = t.get("direction_name", "Общее")
        contractor = t.get("contractor_name", "Без контрагента")
        date = t["date"]
        amount = t["amount"]
        
        # By category
        if cat_name not in by_category:
            by_category[cat_name] = {"amount": 0, "count": 0}
        by_category[cat_name]["amount"] += amount
        by_category[cat_name]["count"] += 1
        
        # By direction
        if dir_name not in by_direction:
            by_direction[dir_name] = {"amount": 0, "count": 0}
        by_direction[dir_name]["amount"] += amount
        by_direction[dir_name]["count"] += 1
        
        # By contractor
        if contractor not in by_contractor:
            by_contractor[contractor] = {"amount": 0, "count": 0}
        by_contractor[contractor]["amount"] += amount
        by_contractor[contractor]["count"] += 1
        
        # Daily
        if date not in daily_expenses:
            daily_expenses[date] = 0
        daily_expenses[date] += amount
    
    total_expense = sum(t["amount"] for t in transactions)
    
    # Sort and limit
    top_categories = sorted(by_category.items(), key=lambda x: x[1]["amount"], reverse=True)[:15]
    top_contractors = sorted(by_contractor.items(), key=lambda x: x[1]["amount"], reverse=True)[:10]
    
    # Calculate average
    days = (datetime.strptime(date_to, "%Y-%m-%d") - datetime.strptime(date_from, "%Y-%m-%d")).days + 1
    daily_average = total_expense / max(days, 1)
    
    return {
        "period": {"from": date_from, "to": date_to},
        "total_expense": total_expense,
        "daily_average": daily_average,
        "transaction_count": len(transactions),
        "by_category": [{"name": k, **v, "percent": (v["amount"]/total_expense*100) if total_expense > 0 else 0} for k, v in top_categories],
        "by_direction": [{"name": k, **v, "percent": (v["amount"]/total_expense*100) if total_expense > 0 else 0} for k, v in by_direction.items()],
        "top_contractors": [{"name": k, **v} for k, v in top_contractors],
        "daily_trend": [{"date": k, "amount": v} for k, v in sorted(daily_expenses.items())]
    }

@api_router.get("/analytics/profitability")
async def get_profitability_report(
    date_from: str,
    date_to: str,
    current_user: dict = Depends(get_current_user)
):
    """Profitability analysis by business direction"""
    query = {
        "user_id": current_user["user_id"],
        "status": "fact",
        "date": {"$gte": date_from, "$lte": date_to}
    }
    
    transactions = await db.transactions.find(query, {"_id": 0}).to_list(10000)
    directions = await db.directions.find({"user_id": current_user["user_id"], "is_active": True}, {"_id": 0}).to_list(20)
    
    # Calculate by direction
    by_direction = {}
    for d in directions:
        by_direction[d["name"]] = {
            "id": d["id"],
            "color": d.get("color", "gray"),
            "income": 0,
            "expense": 0,
            "profit": 0,
            "margin": 0,
            "transactions": 0
        }
    
    for t in transactions:
        dir_name = t.get("direction_name", "Общее")
        if dir_name not in by_direction:
            by_direction[dir_name] = {
                "id": None,
                "color": "gray",
                "income": 0,
                "expense": 0,
                "profit": 0,
                "margin": 0,
                "transactions": 0
            }
        
        by_direction[dir_name]["transactions"] += 1
        
        if t["type"] == "income":
            by_direction[dir_name]["income"] += t["amount"]
        elif t["type"] == "expense":
            by_direction[dir_name]["expense"] += t["amount"]
    
    # Calculate profit and margin
    for name, data in by_direction.items():
        data["profit"] = data["income"] - data["expense"]
        data["margin"] = (data["profit"] / data["income"] * 100) if data["income"] > 0 else 0
    
    # Sort by profit
    sorted_directions = sorted(by_direction.items(), key=lambda x: x[1]["profit"], reverse=True)
    
    total_income = sum(d[1]["income"] for d in sorted_directions)
    total_expense = sum(d[1]["expense"] for d in sorted_directions)
    total_profit = total_income - total_expense
    overall_margin = (total_profit / total_income * 100) if total_income > 0 else 0
    
    return {
        "period": {"from": date_from, "to": date_to},
        "by_direction": [{"name": k, **v} for k, v in sorted_directions],
        "totals": {
            "income": total_income,
            "expense": total_expense,
            "profit": total_profit,
            "margin": overall_margin
        }
    }

@api_router.get("/analytics/top-contractors")
async def get_top_contractors(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 10,
    current_user: dict = Depends(get_current_user)
):
    """Get top contractors by transaction volume"""
    if not date_from:
        date_from = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%d")
    if not date_to:
        date_to = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    query = {
        "user_id": current_user["user_id"],
        "status": "fact",
        "date": {"$gte": date_from, "$lte": date_to},
        "contractor_id": {"$ne": None}
    }
    
    transactions = await db.transactions.find(query, {"_id": 0}).to_list(10000)
    
    # Aggregate by contractor
    contractor_stats = {}
    for t in transactions:
        contractor_id = t.get("contractor_id")
        contractor_name = t.get("contractor_name", "Неизвестный")
        
        if not contractor_id:
            continue
        
        if contractor_id not in contractor_stats:
            contractor_stats[contractor_id] = {
                "id": contractor_id,
                "name": contractor_name,
                "income": 0,
                "expense": 0,
                "total": 0,
                "transactions": 0
            }
        
        contractor_stats[contractor_id]["transactions"] += 1
        
        if t["type"] == "income":
            contractor_stats[contractor_id]["income"] += t["amount"]
            contractor_stats[contractor_id]["total"] += t["amount"]
        elif t["type"] == "expense":
            contractor_stats[contractor_id]["expense"] += t["amount"]
            contractor_stats[contractor_id]["total"] += t["amount"]
    
    # Sort by total volume
    sorted_contractors = sorted(contractor_stats.values(), key=lambda x: x["total"], reverse=True)[:limit]
    
    return {
        "period": {"from": date_from, "to": date_to},
        "contractors": sorted_contractors
    }

# ============== HEALTH CHECK ==============

@api_router.get("/")
async def root():
    return {"message": "WM Finance API", "version": "1.0.0"}

@api_router.get("/health")
async def health():
    return {"status": "healthy"}

# ============== TELEGRAM BOT - ENHANCED SUMMARY ==============

@api_router.get("/bot/summary")
async def get_telegram_summary(
    user_token: str,
    period: Literal["day", "week", "month"] = "week"
):
    """Get AI-powered financial summary for Telegram bot"""
    try:
        payload = jwt.decode(user_token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload["user_id"]
    except (jwt.ExpiredSignatureError, jwt.InvalidTokenError):
        raise HTTPException(status_code=401, detail="Invalid token")
    
    now = datetime.now(timezone.utc)
    
    if period == "day":
        date_from = now.strftime("%Y-%m-%d")
        period_label = "за сегодня"
    elif period == "week":
        date_from = (now - timedelta(days=7)).strftime("%Y-%m-%d")
        period_label = "за неделю"
    else:
        date_from = (now - timedelta(days=30)).strftime("%Y-%m-%d")
        period_label = "за месяц"
    
    date_to = now.strftime("%Y-%m-%d")
    
    # Get transactions
    transactions = await db.transactions.find(
        {"user_id": user_id, "status": "fact", "date": {"$gte": date_from, "$lte": date_to}},
        {"_id": 0}
    ).to_list(10000)
    
    # Calculate totals
    income = sum(t["amount"] for t in transactions if t["type"] == "income")
    expense = sum(t["amount"] for t in transactions if t["type"] == "expense")
    profit = income - expense
    
    # By direction
    by_direction = {}
    for t in transactions:
        dir_name = t.get("direction_name", "Общее")
        if dir_name not in by_direction:
            by_direction[dir_name] = {"income": 0, "expense": 0}
        if t["type"] == "income":
            by_direction[dir_name]["income"] += t["amount"]
        elif t["type"] == "expense":
            by_direction[dir_name]["expense"] += t["amount"]
    
    # Top expense categories
    expense_by_cat = {}
    for t in transactions:
        if t["type"] == "expense":
            cat_name = t.get("category_name", "Прочее")
            expense_by_cat[cat_name] = expense_by_cat.get(cat_name, 0) + t["amount"]
    
    top_expenses = sorted(expense_by_cat.items(), key=lambda x: x[1], reverse=True)[:5]
    
    # Get account balances
    accounts = await db.accounts.find(
        {"user_id": user_id, "is_active": True},
        {"_id": 0}
    ).to_list(20)
    
    total_balance = sum(a.get("current_balance", 0) for a in accounts)
    
    # Get upcoming payments
    upcoming = await db.planned_payments.find(
        {
            "user_id": user_id,
            "status": "pending",
            "date": {"$gte": date_to, "$lte": (now + timedelta(days=7)).strftime("%Y-%m-%d")}
        },
        {"_id": 0}
    ).to_list(10)
    
    upcoming_expense = sum(p["amount"] for p in upcoming if p["type"] == "expense")
    upcoming_income = sum(p["amount"] for p in upcoming if p["type"] == "income")
    
    # Format message
    emoji_profit = "📈" if profit >= 0 else "📉"
    
    message = f"""📊 *Финансовая сводка {period_label}*

💰 *Общие показатели:*
• Доходы: +{income:,.0f} zł
• Расходы: -{expense:,.0f} zł
• {emoji_profit} Прибыль: {profit:,.0f} zł

🏦 *Баланс на счетах:* {total_balance:,.0f} zł

"""
    
    if by_direction:
        message += "📂 *По направлениям:*\n"
        for dir_name, data in by_direction.items():
            dir_profit = data["income"] - data["expense"]
            message += f"• {dir_name}: {dir_profit:+,.0f} zł\n"
        message += "\n"
    
    if top_expenses:
        message += "📌 *Топ расходов:*\n"
        for cat, amount in top_expenses:
            message += f"• {cat}: {amount:,.0f} zł\n"
        message += "\n"
    
    if upcoming_expense > 0 or upcoming_income > 0:
        message += "⏰ *На следующей неделе:*\n"
        if upcoming_income > 0:
            message += f"• Ожидается: +{upcoming_income:,.0f} zł\n"
        if upcoming_expense > 0:
            message += f"• К оплате: -{upcoming_expense:,.0f} zł\n"
    
    return {
        "message": message,
        "data": {
            "period": period,
            "income": income,
            "expense": expense,
            "profit": profit,
            "balance": total_balance,
            "by_direction": by_direction,
            "top_expenses": top_expenses,
            "upcoming_expense": upcoming_expense,
            "upcoming_income": upcoming_income
        }
    }

# ============== INTEGRATION SETTINGS API ==============

@api_router.get("/settings/integrations")
async def get_integration_settings(current_user: dict = Depends(get_current_user)):
    """Get integration settings for current user"""
    settings = await db.integration_settings.find_one(
        {"user_id": current_user["user_id"]},
        {"_id": 0}
    )
    
    if not settings:
        # Create default settings
        settings = {
            "id": str(uuid.uuid4()),
            "user_id": current_user["user_id"],
            "telegram_bot_token": None,
            "telegram_chat_id": None,
            "telegram_auto_summary": False,
            "telegram_summary_schedule": "weekly",
            "telegram_summary_time": "09:00",
            "adesk_api_token": None,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        await db.integration_settings.insert_one(settings)
    
    # Mask sensitive tokens
    result = dict(settings)
    if result.get("telegram_bot_token"):
        result["telegram_bot_token"] = result["telegram_bot_token"][:10] + "..." + result["telegram_bot_token"][-4:]
    if result.get("adesk_api_token"):
        result["adesk_api_token"] = result["adesk_api_token"][:10] + "..." + result["adesk_api_token"][-4:]
    
    return result

@api_router.put("/settings/integrations/telegram")
async def update_telegram_settings(
    data: TelegramSettingsUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update Telegram integration settings"""
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    
    if data.telegram_bot_token is not None:
        update_data["telegram_bot_token"] = data.telegram_bot_token
    if data.telegram_chat_id is not None:
        update_data["telegram_chat_id"] = data.telegram_chat_id
    if data.telegram_auto_summary is not None:
        update_data["telegram_auto_summary"] = data.telegram_auto_summary
    if data.telegram_summary_schedule is not None:
        update_data["telegram_summary_schedule"] = data.telegram_summary_schedule
    if data.telegram_summary_time is not None:
        update_data["telegram_summary_time"] = data.telegram_summary_time
    
    await db.integration_settings.update_one(
        {"user_id": current_user["user_id"]},
        {"$set": update_data},
        upsert=True
    )
    
    return {"status": "updated"}

@api_router.post("/settings/telegram/test")
async def test_telegram_connection(
    data: TelegramTestMessage,
    current_user: dict = Depends(get_current_user)
):
    """Test Telegram bot connection by sending a test message"""
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.post(
                f"https://api.telegram.org/bot{data.bot_token}/sendMessage",
                json={
                    "chat_id": data.chat_id,
                    "text": "✅ *WM Finance подключен!*\n\nТестовое сообщение отправлено успешно.",
                    "parse_mode": "Markdown"
                }
            )
            
            if response.status_code == 200:
                return {"status": "success", "message": "Сообщение отправлено"}
            else:
                error = response.json()
                return {"status": "error", "message": error.get("description", "Ошибка отправки")}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@api_router.post("/settings/telegram/send-summary")
async def send_telegram_summary(
    period: Literal["day", "week", "month"] = "week",
    current_user: dict = Depends(get_current_user)
):
    """Manually send financial summary to Telegram"""
    # Get settings
    settings = await db.integration_settings.find_one(
        {"user_id": current_user["user_id"]},
        {"_id": 0}
    )
    
    if not settings or not settings.get("telegram_bot_token") or not settings.get("telegram_chat_id"):
        raise HTTPException(status_code=400, detail="Telegram не настроен")
    
    # Generate summary message
    now = datetime.now(timezone.utc)
    
    if period == "day":
        date_from = now.strftime("%Y-%m-%d")
        period_label = "за сегодня"
    elif period == "week":
        date_from = (now - timedelta(days=7)).strftime("%Y-%m-%d")
        period_label = "за неделю"
    else:
        date_from = (now - timedelta(days=30)).strftime("%Y-%m-%d")
        period_label = "за месяц"
    
    date_to = now.strftime("%Y-%m-%d")
    
    transactions = await db.transactions.find(
        {"user_id": current_user["user_id"], "status": "fact", "date": {"$gte": date_from, "$lte": date_to}},
        {"_id": 0}
    ).to_list(10000)
    
    income = sum(t["amount"] for t in transactions if t["type"] == "income")
    expense = sum(t["amount"] for t in transactions if t["type"] == "expense")
    profit = income - expense
    
    accounts = await db.accounts.find(
        {"user_id": current_user["user_id"], "is_active": True},
        {"_id": 0}
    ).to_list(20)
    
    total_balance = sum(a.get("current_balance", 0) for a in accounts)
    
    emoji_profit = "📈" if profit >= 0 else "📉"
    
    message = f"""📊 *Финансовая сводка {period_label}*

💰 *Показатели:*
• Доходы: +{income:,.0f} zł
• Расходы: -{expense:,.0f} zł
• {emoji_profit} Прибыль: {profit:,.0f} zł

🏦 *Баланс:* {total_balance:,.0f} zł

_Отправлено из WM Finance_"""
    
    # Send to Telegram
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.post(
                f"https://api.telegram.org/bot{settings['telegram_bot_token']}/sendMessage",
                json={
                    "chat_id": settings["telegram_chat_id"],
                    "text": message,
                    "parse_mode": "Markdown"
                }
            )
            
            if response.status_code == 200:
                return {"status": "success", "message": "Сводка отправлена в Telegram"}
            else:
                error = response.json()
                raise HTTPException(status_code=400, detail=error.get("description", "Ошибка отправки"))
    except httpx.TimeoutException:
        raise HTTPException(status_code=500, detail="Таймаут подключения к Telegram")

# ============== ADESK MIGRATION API ==============

import httpx

@api_router.post("/adesk/test-connection")
async def test_adesk_connection(
    data: AdeskConnectionTest,
    current_user: dict = Depends(get_current_user)
):
    """Test connection to Adesk API"""
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            # Adesk API uses token as query parameter, not header
            params = {
                "api_token": data.api_token,
                "limit": 100,
                "startDate": "01.01.2020",
                "endDate": "31.12.2026"
            }
            
            response = await client.get(
                "https://api.adesk.ru/v1/transactions",
                params=params
            )
            logger.info(f"Adesk API response: status={response.status_code}")
            
            if response.status_code == 200:
                data_response = response.json()
                logger.info(f"Adesk response type: {type(data_response)}, keys: {data_response.keys() if isinstance(data_response, dict) else 'list'}")
                
                # Adesk returns data in different formats
                if isinstance(data_response, list):
                    transactions_count = len(data_response)
                elif isinstance(data_response, dict):
                    transactions_count = len(data_response.get("data", data_response.get("items", data_response.get("transactions", []))))
                else:
                    transactions_count = 0
                
                return {
                    "status": "success", 
                    "message": "Подключение успешно",
                    "transactions_count": transactions_count
                }
            elif response.status_code == 401 or response.status_code == 403:
                return {"status": "error", "message": "Неверный API токен"}
            else:
                logger.error(f"Adesk API error: {response.status_code} - {response.text[:500]}")
                return {"status": "error", "message": f"Ошибка API: {response.status_code}"}
                
    except httpx.TimeoutException:
        return {"status": "error", "message": "Таймаут подключения к Adesk"}
    except Exception as e:
        logger.error(f"Adesk connection error: {e}")
        return {"status": "error", "message": f"Ошибка подключения: {str(e)}"}

@api_router.post("/adesk/start-migration")
async def start_adesk_migration(
    data: AdeskMigrationStart,
    current_user: dict = Depends(get_current_user)
):
    """Start migration from Adesk - load data into drafts"""
    batch_id = str(uuid.uuid4())
    
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            # Get existing mappings
            categories = await db.categories.find({"user_id": current_user["user_id"]}, {"_id": 0}).to_list(200)
            directions = await db.directions.find({"user_id": current_user["user_id"]}, {"_id": 0}).to_list(20)
            contractors = await db.contractors.find({"user_id": current_user["user_id"]}, {"_id": 0}).to_list(500)
            accounts = await db.accounts.find({"user_id": current_user["user_id"]}, {"_id": 0}).to_list(50)
            
            # Create lookup maps
            category_map = {c["name"].lower(): c for c in categories}
            direction_map = {d["name"].lower(): d for d in directions}
            contractor_map = {c["name"].lower(): c for c in contractors}
            account_map = {a["name"].lower(): a for a in accounts}
            
            # Smart mapping for projects to directions
            project_direction_map = {
                "теплиц": "теплицы",
                "саун": "сауны",
                "купел": "купели",
                "бан": "сауны",
            }
            
            drafts_created = 0
            errors = 0
            
            # Fetch transactions from Adesk - token as query param
            if data.migrate_transactions:
                page = 1
                while True:
                    # Convert dates to Adesk format (DD.MM.YYYY)
                    start_date = data.date_from.replace("-", ".") if "-" in data.date_from else data.date_from
                    end_date = data.date_to.replace("-", ".") if "-" in data.date_to else data.date_to
                    
                    # Reformat if needed (YYYY-MM-DD -> DD.MM.YYYY)
                    if len(start_date) == 10 and start_date[4] == ".":
                        parts = start_date.split(".")
                        start_date = f"{parts[2]}.{parts[1]}.{parts[0]}"
                    if len(end_date) == 10 and end_date[4] == ".":
                        parts = end_date.split(".")
                        end_date = f"{parts[2]}.{parts[1]}.{parts[0]}"
                    
                    response = await client.get(
                        "https://api.adesk.ru/v1/transactions",
                        params={
                            "api_token": data.api_token,
                            "startDate": start_date,
                            "endDate": end_date,
                            "limit": 100,
                            "page": page
                        }
                    )
                    
                    logger.info(f"Adesk migration page {page}: status={response.status_code}")
                    
                    if response.status_code != 200:
                        logger.error(f"Adesk API error: {response.text[:500]}")
                        break
                    
                    result = response.json()
                    
                    # Handle different response formats
                    if isinstance(result, list):
                        transactions = result
                    elif isinstance(result, dict):
                        transactions = result.get("data", result.get("items", result.get("transactions", [])))
                    else:
                        transactions = []
                    
                    logger.info(f"Adesk page {page}: found {len(transactions)} transactions")
                    
                    if not transactions:
                        break
                    
                    for t in transactions:
                        try:
                            # Determine type
                            t_type = "expense"
                            if t.get("type") == "income" or t.get("is_income"):
                                t_type = "income"
                            elif t.get("type") == "transfer":
                                t_type = "transfer"
                            
                            # Smart mapping
                            status = "ready"
                            error_reason = None
                            
                            # Map category
                            cat_adesk = t.get("category", {}).get("name", "") or t.get("category_name", "")
                            mapped_cat = category_map.get(cat_adesk.lower())
                            
                            # Map direction from project
                            project_adesk = t.get("project", {}).get("name", "") or t.get("project_name", "")
                            mapped_dir = None
                            for key, dir_name in project_direction_map.items():
                                if key in project_adesk.lower():
                                    mapped_dir = direction_map.get(dir_name)
                                    break
                            
                            # Map contractor
                            contractor_adesk = t.get("contractor", {}).get("name", "") or t.get("contractor_name", "")
                            mapped_contractor = contractor_map.get(contractor_adesk.lower())
                            
                            # Map account
                            account_adesk = t.get("account", {}).get("name", "") or t.get("account_name", "")
                            mapped_account = account_map.get(account_adesk.lower())
                            
                            # Determine status
                            if not mapped_cat:
                                status = "needs_review"
                            if not mapped_dir:
                                status = "needs_review"
                            if not mapped_account:
                                status = "needs_review"
                            
                            draft = {
                                "id": str(uuid.uuid4()),
                                "created_at": datetime.now(timezone.utc).isoformat(),
                                "adesk_id": str(t.get("id", "")),
                                "date": t.get("date", "")[:10] if t.get("date") else data.date_from,
                                "type": t_type,
                                "amount": abs(float(t.get("amount", 0))),
                                "currency": t.get("currency", "PLN"),
                                "category_adesk": cat_adesk,
                                "category_id": mapped_cat["id"] if mapped_cat else None,
                                "category_name": mapped_cat["name"] if mapped_cat else None,
                                "project_adesk": project_adesk,
                                "direction_id": mapped_dir["id"] if mapped_dir else None,
                                "direction_name": mapped_dir["name"] if mapped_dir else None,
                                "contractor_adesk": contractor_adesk,
                                "contractor_id": mapped_contractor["id"] if mapped_contractor else None,
                                "contractor_name": mapped_contractor["name"] if mapped_contractor else None,
                                "account_adesk": account_adesk,
                                "account_id": mapped_account["id"] if mapped_account else None,
                                "account_name": mapped_account["name"] if mapped_account else None,
                                "description": t.get("description", "") or t.get("comment", ""),
                                "status": status,
                                "error_reason": error_reason,
                                "user_id": current_user["user_id"],
                                "batch_id": batch_id
                            }
                            
                            await db.adesk_drafts.insert_one(draft)
                            drafts_created += 1
                            
                        except Exception as e:
                            logger.error(f"Error processing Adesk transaction: {e}")
                            errors += 1
                    
                    page += 1
                    if len(transactions) < 100:
                        break
            
            # Get stats
            ready_count = await db.adesk_drafts.count_documents(
                {"batch_id": batch_id, "status": "ready"}
            )
            review_count = await db.adesk_drafts.count_documents(
                {"batch_id": batch_id, "status": "needs_review"}
            )
            
            return {
                "status": "success",
                "batch_id": batch_id,
                "drafts_created": drafts_created,
                "ready": ready_count,
                "needs_review": review_count,
                "errors": errors
            }
            
    except Exception as e:
        logger.error(f"Migration error: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка миграции: {str(e)}")

@api_router.get("/adesk/drafts")
async def get_adesk_drafts(
    batch_id: Optional[str] = None,
    status: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    current_user: dict = Depends(get_current_user)
):
    """Get Adesk migration drafts"""
    query = {"user_id": current_user["user_id"]}
    
    if batch_id:
        query["batch_id"] = batch_id
    if status:
        query["status"] = status
    
    total = await db.adesk_drafts.count_documents(query)
    
    drafts = await db.adesk_drafts.find(query, {"_id": 0}).sort("date", -1).skip((page - 1) * limit).limit(limit).to_list(limit)
    
    # Get stats
    stats = {
        "total": total,
        "ready": await db.adesk_drafts.count_documents({**query, "status": "ready"}),
        "needs_review": await db.adesk_drafts.count_documents({**query, "status": "needs_review"}),
        "error": await db.adesk_drafts.count_documents({**query, "status": "error"}),
        "imported": await db.adesk_drafts.count_documents({**query, "status": "imported"})
    }
    
    return {
        "drafts": drafts,
        "stats": stats,
        "page": page,
        "limit": limit,
        "total": total
    }

@api_router.put("/adesk/drafts/{draft_id}")
async def update_adesk_draft(
    draft_id: str,
    data: AdeskDraftUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update a single Adesk draft"""
    draft = await db.adesk_drafts.find_one(
        {"id": draft_id, "user_id": current_user["user_id"]},
        {"_id": 0}
    )
    
    if not draft:
        raise HTTPException(status_code=404, detail="Draft not found")
    
    update_data = {}
    
    if data.category_id:
        cat = await db.categories.find_one({"id": data.category_id}, {"_id": 0})
        if cat:
            update_data["category_id"] = cat["id"]
            update_data["category_name"] = cat["name"]
    
    if data.direction_id:
        dir_obj = await db.directions.find_one({"id": data.direction_id}, {"_id": 0})
        if dir_obj:
            update_data["direction_id"] = dir_obj["id"]
            update_data["direction_name"] = dir_obj["name"]
    
    if data.contractor_id:
        contractor = await db.contractors.find_one({"id": data.contractor_id}, {"_id": 0})
        if contractor:
            update_data["contractor_id"] = contractor["id"]
            update_data["contractor_name"] = contractor["name"]
    
    if data.account_id:
        account = await db.accounts.find_one({"id": data.account_id}, {"_id": 0})
        if account:
            update_data["account_id"] = account["id"]
            update_data["account_name"] = account["name"]
    
    if data.description is not None:
        update_data["description"] = data.description
    
    # Check if ready
    draft_updated = {**draft, **update_data}
    if draft_updated.get("category_id") and draft_updated.get("direction_id") and draft_updated.get("account_id"):
        update_data["status"] = "ready"
    
    await db.adesk_drafts.update_one(
        {"id": draft_id},
        {"$set": update_data}
    )
    
    return {"status": "updated"}

@api_router.post("/adesk/drafts/bulk-update")
async def bulk_update_adesk_drafts(
    data: AdeskBulkUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Bulk update multiple drafts"""
    update_data = {}
    
    if data.category_id:
        cat = await db.categories.find_one({"id": data.category_id}, {"_id": 0})
        if cat:
            update_data["category_id"] = cat["id"]
            update_data["category_name"] = cat["name"]
    
    if data.direction_id:
        dir_obj = await db.directions.find_one({"id": data.direction_id}, {"_id": 0})
        if dir_obj:
            update_data["direction_id"] = dir_obj["id"]
            update_data["direction_name"] = dir_obj["name"]
    
    if data.contractor_id:
        contractor = await db.contractors.find_one({"id": data.contractor_id}, {"_id": 0})
        if contractor:
            update_data["contractor_id"] = contractor["id"]
            update_data["contractor_name"] = contractor["name"]
    
    if update_data:
        await db.adesk_drafts.update_many(
            {"id": {"$in": data.draft_ids}, "user_id": current_user["user_id"]},
            {"$set": update_data}
        )
    
    return {"status": "updated", "count": len(data.draft_ids)}

@api_router.post("/adesk/confirm-ready")
async def confirm_ready_drafts(
    batch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Confirm and import all ready drafts into main database"""
    query = {"user_id": current_user["user_id"], "status": "ready"}
    if batch_id:
        query["batch_id"] = batch_id
    
    ready_drafts = await db.adesk_drafts.find(query, {"_id": 0}).to_list(10000)
    
    imported = 0
    duplicates = 0
    errors = 0
    
    for draft in ready_drafts:
        try:
            # Check for duplicate
            existing = await db.transactions.find_one({
                "user_id": current_user["user_id"],
                "date": draft["date"],
                "amount": draft["amount"],
                "account_id": draft["account_id"]
            })
            
            if existing:
                await db.adesk_drafts.update_one(
                    {"id": draft["id"]},
                    {"$set": {"status": "error", "error_reason": "Дубликат операции"}}
                )
                duplicates += 1
                continue
            
            # Create transaction
            transaction = {
                "id": str(uuid.uuid4()),
                "date": draft["date"],
                "type": draft["type"],
                "amount": draft["amount"],
                "currency": draft["currency"],
                "category_id": draft["category_id"],
                "category_name": draft["category_name"],
                "direction_id": draft["direction_id"],
                "direction_name": draft["direction_name"],
                "account_id": draft["account_id"],
                "account_name": draft["account_name"],
                "contractor_id": draft["contractor_id"],
                "contractor_name": draft["contractor_name"],
                "description": draft["description"],
                "status": "fact",
                "source": "adesk_migration",
                "adesk_id": draft["adesk_id"],
                "balance_after": 0,
                "user_id": current_user["user_id"],
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.transactions.insert_one(transaction)
            
            # Update account balance
            if draft["account_id"]:
                balance_change = draft["amount"] if draft["type"] == "income" else -draft["amount"]
                await db.accounts.update_one(
                    {"id": draft["account_id"]},
                    {"$inc": {"current_balance": balance_change}}
                )
            
            # Mark draft as imported
            await db.adesk_drafts.update_one(
                {"id": draft["id"]},
                {"$set": {"status": "imported"}}
            )
            
            imported += 1
            
        except Exception as e:
            logger.error(f"Error importing draft {draft['id']}: {e}")
            errors += 1
    
    return {
        "status": "success",
        "imported": imported,
        "duplicates": duplicates,
        "errors": errors
    }

@api_router.delete("/adesk/drafts/{draft_id}")
async def delete_adesk_draft(
    draft_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a single draft"""
    await db.adesk_drafts.delete_one(
        {"id": draft_id, "user_id": current_user["user_id"]}
    )
    return {"status": "deleted"}

@api_router.delete("/adesk/drafts")
async def delete_all_drafts(
    batch_id: Optional[str] = None,
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Delete drafts by filter"""
    query = {"user_id": current_user["user_id"]}
    if batch_id:
        query["batch_id"] = batch_id
    if status:
        query["status"] = status
    
    result = await db.adesk_drafts.delete_many(query)
    return {"status": "deleted", "count": result.deleted_count}

@api_router.get("/adesk/export-problems")
async def export_problem_drafts(
    batch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export problematic drafts as CSV"""
    query = {
        "user_id": current_user["user_id"],
        "status": {"$in": ["needs_review", "error"]}
    }
    if batch_id:
        query["batch_id"] = batch_id
    
    drafts = await db.adesk_drafts.find(query, {"_id": 0}).to_list(10000)
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Дата", "Тип", "Сумма", "Валюта", "Категория Adesk", "Категория WM",
        "Проект Adesk", "Направление WM", "Контрагент Adesk", "Контрагент WM",
        "Счёт Adesk", "Счёт WM", "Описание", "Статус", "Причина ошибки"
    ])
    
    for d in drafts:
        writer.writerow([
            d.get("date"), d.get("type"), d.get("amount"), d.get("currency"),
            d.get("category_adesk"), d.get("category_name") or "-",
            d.get("project_adesk"), d.get("direction_name") or "-",
            d.get("contractor_adesk"), d.get("contractor_name") or "-",
            d.get("account_adesk"), d.get("account_name") or "-",
            d.get("description"), d.get("status"), d.get("error_reason") or "-"
        ])
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=adesk_problems.csv"}
    )

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============== GOOGLE SHEETS BACKUP ==============

GOOGLE_SHEETS_SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]
GOOGLE_SERVICE_ACCOUNT_FILE = ROOT_DIR / 'google_service_account.json'
GOOGLE_SPREADSHEET_ID = "1unqs6afSAc3V-M7f3keGh_M6gNpo-VpmUj50nZZKY3Q"

def get_gspread_client():
    """Get authenticated gspread client"""
    if not GOOGLE_SERVICE_ACCOUNT_FILE.exists():
        logger.error("Google service account file not found")
        return None
    
    creds = ServiceAccountCredentials.from_service_account_file(
        str(GOOGLE_SERVICE_ACCOUNT_FILE),
        scopes=GOOGLE_SHEETS_SCOPES
    )
    return gspread.authorize(creds)

async def backup_to_google_sheets(user_id: str = None):
    """Export all data to Google Sheets"""
    logger.info("Starting Google Sheets backup...")
    
    try:
        gc = get_gspread_client()
        if not gc:
            logger.error("Failed to get gspread client")
            return {"status": "error", "message": "Google Sheets не настроен"}
        
        spreadsheet = gc.open_by_key(GOOGLE_SPREADSHEET_ID)
        
        # Query filter - if user_id provided, filter by user
        query = {"user_id": user_id} if user_id else {}
        
        now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        
        # 1. TRANSACTIONS
        transactions = await db.transactions.find(query, {"_id": 0}).sort("date", -1).to_list(50000)
        trans_headers = ["ID", "Дата", "Тип", "Сумма", "Валюта", "Категория", "Направление", "Счёт", "Контрагент", "Описание", "Статус", "Источник"]
        trans_rows = [trans_headers]
        for t in transactions:
            trans_rows.append([
                t.get("id", ""),
                t.get("date", ""),
                t.get("type", ""),
                t.get("amount", 0),
                t.get("currency", "PLN"),
                t.get("category_name", ""),
                t.get("direction_name", ""),
                t.get("account_name", ""),
                t.get("contractor_name", ""),
                t.get("description", ""),
                t.get("status", ""),
                t.get("source", "")
            ])
        
        # Get or create worksheet
        try:
            ws_trans = spreadsheet.worksheet("Операции")
            ws_trans.clear()
        except gspread.WorksheetNotFound:
            ws_trans = spreadsheet.add_worksheet("Операции", rows=len(trans_rows)+10, cols=15)
        
        if trans_rows:
            ws_trans.update(trans_rows, value_input_option='RAW')
        
        # 2. CONTRACTORS
        contractors = await db.contractors.find({**query, "is_active": True}, {"_id": 0}).to_list(5000)
        contr_headers = ["ID", "Название", "Тип", "Группа", "Email", "Телефон", "Комментарий"]
        contr_rows = [contr_headers]
        for c in contractors:
            contr_rows.append([
                c.get("id", ""),
                c.get("name", ""),
                c.get("type", ""),
                c.get("group", ""),
                c.get("email", ""),
                c.get("phone", ""),
                c.get("comment", "")
            ])
        
        try:
            ws_contr = spreadsheet.worksheet("Контрагенты")
            ws_contr.clear()
        except gspread.WorksheetNotFound:
            ws_contr = spreadsheet.add_worksheet("Контрагенты", rows=len(contr_rows)+10, cols=10)
        
        if contr_rows:
            ws_contr.update(contr_rows, value_input_option='RAW')
        
        # 3. ACCOUNTS
        accounts = await db.accounts.find({**query, "is_active": True}, {"_id": 0}).to_list(100)
        acc_headers = ["ID", "Название", "Тип", "Валюта", "Банк", "Начальный баланс", "Текущий баланс"]
        acc_rows = [acc_headers]
        for a in accounts:
            acc_rows.append([
                a.get("id", ""),
                a.get("name", ""),
                a.get("type", ""),
                a.get("currency", ""),
                a.get("bank", ""),
                a.get("initial_balance", 0),
                a.get("current_balance", 0)
            ])
        
        try:
            ws_acc = spreadsheet.worksheet("Счета")
            ws_acc.clear()
        except gspread.WorksheetNotFound:
            ws_acc = spreadsheet.add_worksheet("Счета", rows=len(acc_rows)+10, cols=10)
        
        if acc_rows:
            ws_acc.update(acc_rows, value_input_option='RAW')
        
        # 4. CATEGORIES
        categories = await db.categories.find({**query, "is_active": True}, {"_id": 0}).to_list(500)
        cat_headers = ["ID", "Название", "Тип", "Группа"]
        cat_rows = [cat_headers]
        for c in categories:
            cat_rows.append([
                c.get("id", ""),
                c.get("name", ""),
                c.get("type", ""),
                c.get("group", "")
            ])
        
        try:
            ws_cat = spreadsheet.worksheet("Категории")
            ws_cat.clear()
        except gspread.WorksheetNotFound:
            ws_cat = spreadsheet.add_worksheet("Категории", rows=len(cat_rows)+10, cols=6)
        
        if cat_rows:
            ws_cat.update(cat_rows, value_input_option='RAW')
        
        # 5. DIRECTIONS
        directions = await db.directions.find({**query, "is_active": True}, {"_id": 0}).to_list(50)
        dir_headers = ["ID", "Название", "Цвет", "Описание"]
        dir_rows = [dir_headers]
        for d in directions:
            dir_rows.append([
                d.get("id", ""),
                d.get("name", ""),
                d.get("color", ""),
                d.get("description", "")
            ])
        
        try:
            ws_dir = spreadsheet.worksheet("Направления")
            ws_dir.clear()
        except gspread.WorksheetNotFound:
            ws_dir = spreadsheet.add_worksheet("Направления", rows=len(dir_rows)+10, cols=6)
        
        if dir_rows:
            ws_dir.update(dir_rows, value_input_option='RAW')
        
        # 6. PROJECTS
        projects = await db.projects.find(query, {"_id": 0}).to_list(5000)
        proj_headers = ["ID", "Название", "Направление", "Контрагент", "План", "Факт", "Статус", "Начало", "Окончание"]
        proj_rows = [proj_headers]
        for p in projects:
            proj_rows.append([
                p.get("id", ""),
                p.get("name", ""),
                p.get("direction_name", ""),
                p.get("contractor_name", ""),
                p.get("planned_amount", 0),
                p.get("actual_amount", 0),
                p.get("status", ""),
                p.get("start_date", ""),
                p.get("end_date", "")
            ])
        
        try:
            ws_proj = spreadsheet.worksheet("Проекты")
            ws_proj.clear()
        except gspread.WorksheetNotFound:
            ws_proj = spreadsheet.add_worksheet("Проекты", rows=len(proj_rows)+10, cols=12)
        
        if proj_rows:
            ws_proj.update(proj_rows, value_input_option='RAW')
        
        # 7. PLANNED PAYMENTS
        planned = await db.planned_payments.find(query, {"_id": 0}).to_list(5000)
        plan_headers = ["ID", "Дата", "Тип", "Сумма", "Валюта", "Категория", "Контрагент", "Направление", "Счёт", "Статус", "Повтор", "Комментарий"]
        plan_rows = [plan_headers]
        for p in planned:
            plan_rows.append([
                p.get("id", ""),
                p.get("date", ""),
                p.get("type", ""),
                p.get("amount", 0),
                p.get("currency", "PLN"),
                p.get("category_name", ""),
                p.get("contractor_name", ""),
                p.get("direction_name", ""),
                p.get("account_name", ""),
                p.get("status", ""),
                p.get("recurrence", ""),
                p.get("comment", "")
            ])
        
        try:
            ws_plan = spreadsheet.worksheet("Плановые платежи")
            ws_plan.clear()
        except gspread.WorksheetNotFound:
            ws_plan = spreadsheet.add_worksheet("Плановые платежи", rows=len(plan_rows)+10, cols=15)
        
        if plan_rows:
            ws_plan.update(plan_rows, value_input_option='RAW')
        
        # 8. INFO SHEET
        info_rows = [
            ["WM Finance - Автоматический бэкап"],
            [""],
            ["Последнее обновление:", now],
            [""],
            ["Статистика:"],
            ["Операций:", len(transactions)],
            ["Контрагентов:", len(contractors)],
            ["Счетов:", len(accounts)],
            ["Категорий:", len(categories)],
            ["Направлений:", len(directions)],
            ["Проектов:", len(projects)],
            ["Плановых платежей:", len(planned)]
        ]
        
        try:
            ws_info = spreadsheet.worksheet("Инфо")
            ws_info.clear()
        except gspread.WorksheetNotFound:
            ws_info = spreadsheet.add_worksheet("Инфо", rows=20, cols=5)
        
        ws_info.update(info_rows, value_input_option='RAW')
        
        logger.info(f"Google Sheets backup completed: {len(transactions)} transactions, {len(contractors)} contractors")
        
        return {
            "status": "success",
            "message": "Бэкап выполнен",
            "timestamp": now,
            "stats": {
                "transactions": len(transactions),
                "contractors": len(contractors),
                "accounts": len(accounts),
                "categories": len(categories),
                "directions": len(directions),
                "projects": len(projects),
                "planned_payments": len(planned)
            }
        }
        
    except Exception as e:
        logger.error(f"Google Sheets backup error: {e}")
        return {"status": "error", "message": str(e)}

async def scheduled_google_sheets_backup():
    """Scheduled job for daily Google Sheets backup"""
    logger.info("Running scheduled Google Sheets backup")
    result = await backup_to_google_sheets()
    logger.info(f"Scheduled backup result: {result.get('status')}")

@app.post("/api/backup/google-sheets")
async def trigger_backup(current_user: dict = Depends(get_current_user)):
    """Manually trigger Google Sheets backup"""
    result = await backup_to_google_sheets(current_user["user_id"])
    return result

@app.get("/api/backup/status")
async def get_backup_status(current_user: dict = Depends(get_current_user)):
    """Get backup configuration status"""
    gc = get_gspread_client()
    if not gc:
        return {
            "configured": False,
            "message": "Google Sheets не настроен"
        }
    
    try:
        spreadsheet = gc.open_by_key(GOOGLE_SPREADSHEET_ID)
        return {
            "configured": True,
            "spreadsheet_id": GOOGLE_SPREADSHEET_ID,
            "spreadsheet_title": spreadsheet.title,
            "spreadsheet_url": f"https://docs.google.com/spreadsheets/d/{GOOGLE_SPREADSHEET_ID}"
        }
    except Exception as e:
        return {
            "configured": False,
            "message": str(e)
        }

# ============== TELEGRAM SCHEDULER ==============

scheduler = AsyncIOScheduler()

async def send_scheduled_telegram_summary():
    """Background job to send automatic Telegram summaries"""
    logger.info("Running scheduled Telegram summary job")
    
    try:
        # Find all users with auto-summary enabled
        all_settings = await db.integration_settings.find(
            {"telegram_auto_summary": True},
            {"_id": 0}
        ).to_list(1000)
        
        for settings in all_settings:
            if not settings.get("telegram_bot_token") or not settings.get("telegram_chat_id"):
                continue
            
            user_id = settings.get("user_id")
            if not user_id:
                continue
            
            # Check schedule
            schedule = settings.get("telegram_summary_schedule", "weekly")
            now = datetime.now(timezone.utc)
            
            # Validate schedule matches current day
            if schedule == "daily":
                pass  # Always send
            elif schedule == "weekly" and now.weekday() != 0:  # Monday
                continue
            elif schedule == "monday" and now.weekday() != 0:
                continue
            elif schedule == "friday" and now.weekday() != 4:
                continue
            
            try:
                # Calculate period
                if schedule == "daily":
                    period_days = 1
                    period_name = "вчера"
                else:
                    period_days = 7
                    period_name = "неделю"
                
                date_from = (now - timedelta(days=period_days)).strftime("%Y-%m-%d")
                date_to = now.strftime("%Y-%m-%d")
                
                # Get transactions
                transactions = await db.transactions.find({
                    "user_id": user_id,
                    "status": "fact",
                    "date": {"$gte": date_from, "$lte": date_to}
                }, {"_id": 0}).to_list(10000)
                
                total_income = sum(t["amount"] for t in transactions if t["type"] == "income")
                total_expense = sum(t["amount"] for t in transactions if t["type"] == "expense")
                profit = total_income - total_expense
                
                # Get balances
                accounts = await db.accounts.find(
                    {"user_id": user_id, "is_active": True},
                    {"_id": 0}
                ).to_list(100)
                total_balance = sum(a.get("current_balance", 0) for a in accounts)
                
                # Get overdue payments
                overdue = await db.planned_payments.count_documents({
                    "user_id": user_id,
                    "status": "pending",
                    "date": {"$lt": now.strftime("%Y-%m-%d")}
                })
                
                # Build message
                message = f"""📊 *Автоматическая сводка WM Finance*
за {period_name} ({date_from} — {date_to})

💰 Доходы: {total_income:,.2f} zł
💸 Расходы: {total_expense:,.2f} zł
📈 Прибыль: {profit:+,.2f} zł

💳 Остаток на счетах: {total_balance:,.2f} zł"""

                if overdue > 0:
                    message += f"\n\n⚠️ Просроченных платежей: {overdue}"
                
                # Send to Telegram
                import httpx
                async with httpx.AsyncClient(timeout=30.0) as client:
                    await client.post(
                        f"https://api.telegram.org/bot{settings['telegram_bot_token']}/sendMessage",
                        json={
                            "chat_id": settings["telegram_chat_id"],
                            "text": message,
                            "parse_mode": "Markdown"
                        }
                    )
                    logger.info(f"Sent scheduled summary to user {user_id}")
                    
            except Exception as e:
                logger.error(f"Error sending summary for user {user_id}: {e}")
                
    except Exception as e:
        logger.error(f"Scheduler job error: {e}")

def setup_scheduler():
    """Setup the scheduler with jobs"""
    # Run every day at 9:00 UTC (the job itself checks user schedules)
    scheduler.add_job(
        send_scheduled_telegram_summary,
        CronTrigger(hour=9, minute=0),
        id="telegram_summary_job",
        replace_existing=True
    )
    # Google Sheets backup every day at 6:00 UTC (before work day)
    scheduler.add_job(
        scheduled_google_sheets_backup,
        CronTrigger(hour=6, minute=0),
        id="google_sheets_backup_job",
        replace_existing=True
    )
    scheduler.start()
    logger.info("Scheduler started: Telegram summaries at 9:00 UTC, Google Sheets backup at 6:00 UTC")

@app.on_event("startup")
async def startup_event():
    setup_scheduler()

@app.on_event("shutdown")
async def shutdown_db_client():
    scheduler.shutdown(wait=False)
    client.close()
