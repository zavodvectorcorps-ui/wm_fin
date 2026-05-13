from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query
from typing import List, Optional
from datetime import datetime, timezone, timedelta
import csv
import io
import logging

from database import db
from auth import get_current_user
from models import Transaction, TransactionCreate
from services.balance import update_account_balance
from routes.exchange_rate import get_nbp_rate

router = APIRouter(prefix="/api")
logger = logging.getLogger(__name__)


async def get_effective_rate(user_id: str) -> float:
    """Get the effective EUR/PLN rate (manual or NBP)."""
    settings = await db.integration_settings.find_one(
        {"user_id": user_id}, {"_id": 0, "manual_eur_pln_rate": 1}
    )
    manual = settings.get("manual_eur_pln_rate") if settings else None
    if manual:
        return float(manual)
    return await get_nbp_rate()


async def calc_amount_base(amount: float, currency: str, account_id: str, user_id: str):
    """Convert amount to account's currency. Returns (amount_base, exchange_rate)."""
    account = await db.accounts.find_one({"id": account_id, "user_id": user_id}, {"_id": 0, "currency": 1})
    acc_currency = account.get("currency", "PLN") if account else "PLN"

    if currency == acc_currency:
        return amount, None

    rate = await get_effective_rate(user_id)
    if rate <= 0:
        return amount, None

    # EUR -> PLN
    if currency == "EUR" and acc_currency == "PLN":
        return round(amount * rate, 2), rate
    # PLN -> EUR
    if currency == "PLN" and acc_currency == "EUR":
        return round(amount / rate, 2), rate

    return amount, None


@router.get("/transactions")
async def get_transactions(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    type: Optional[str] = None,
    status: Optional[str] = None,
    account_id: Optional[str] = None,
    account_ids: Optional[str] = None,  # comma-separated include list (overrides account_id)
    direction_id: Optional[str] = None,
    category_id: Optional[str] = None,
    contractor_id: Optional[str] = None,
    source: Optional[str] = None,
    search: Optional[str] = None,
    needs_review: Optional[bool] = None,
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    current_user: dict = Depends(get_current_user)
):
    query = {"user_id": current_user["user_id"]}

    if date_from:
        query["date"] = {"$gte": date_from}
    if date_to:
        if "date" in query:
            query["date"]["$lte"] = date_to
        else:
            query["date"] = {"$lte": date_to}
    if type:
        if type == "exchange":
            # UI-only type — actual rows are stored as "transfer" with is_exchange=True
            query["type"] = "transfer"
            query["is_exchange"] = True
        elif type == "transfer":
            # Pure transfer = exclude currency exchanges
            query["type"] = "transfer"
            query["is_exchange"] = {"$ne": True}
        else:
            query["type"] = type
    if status:
        query["status"] = status
    # Multi-account filter (comma-separated). Takes precedence over single account_id.
    if account_ids:
        ids = [s.strip() for s in account_ids.split(",") if s.strip()]
        if ids:
            query["$or"] = [
                {"account_id": {"$in": ids}},
                {"to_account_id": {"$in": ids}},
            ]
    elif account_id:
        query["$or"] = [{"account_id": account_id}, {"to_account_id": account_id}]
    if direction_id:
        query["direction_id"] = direction_id
    if category_id:
        query["category_id"] = category_id
    if contractor_id:
        query["contractor_id"] = contractor_id
    if source:
        query["source"] = source
    if search:
        query["description"] = {"$regex": search, "$options": "i"}
    if needs_review is not None:
        query["needs_review"] = needs_review

    total = await db.transactions.count_documents(query)
    skip = (page - 1) * per_page
    items = await db.transactions.find(query, {"_id": 0}).sort("date", -1).skip(skip).limit(per_page).to_list(per_page)

    # Identify loan accounts so we can split summaries between assets and loans
    loan_accs = await db.accounts.find(
        {"user_id": current_user["user_id"], "is_loan": True},
        {"_id": 0, "id": 1}
    ).to_list(None)
    loan_account_ids = [a["id"] for a in loan_accs]

    # Aggregate summary for ENTIRE filtered period using a COPY of the query
    match_query = {k: v for k, v in query.items()}
    # Exclude TRANSFERS that touch a loan account from the main income/expense
    # summary. Direct income/expense on a loan account remain in the summary —
    # they represent real business cash flow (e.g. paying a supplier from a
    # credit line) and the resulting debt change is already visible in the
    # account balance and in the loans block via `current_balance`.
    main_match = {k: v for k, v in match_query.items()}
    if loan_account_ids:
        main_match["$nor"] = [
            {"type": "transfer", "account_id": {"$in": loan_account_ids}},
            {"type": "transfer", "to_account_id": {"$in": loan_account_ids}},
        ]
    summary_pipeline = [
        {"$match": main_match},
        {"$group": {
            "_id": {
                "currency": {"$ifNull": ["$currency", "PLN"]},
                "type": "$type",
            },
            "total_amount": {"$sum": "$amount"},
            "total_amount_base": {"$sum": {"$ifNull": ["$amount_base", "$amount"]}},
            "count": {"$sum": 1},
        }},
    ]
    summary_raw = await db.transactions.aggregate(summary_pipeline).to_list(100)

    summary = {}
    summary_total_count = 0
    for row in summary_raw:
        cur = row["_id"]["currency"]
        t = row["_id"]["type"]
        # When account filter is active, transfers are handled separately below
        # to properly distinguish direction (in/out) for the filtered account.
        # When no filter, transfers are zero-sum globally — skip them too.
        if t == "transfer":
            summary_total_count += row["count"]
            continue
        if cur not in summary:
            summary[cur] = {"income": 0, "expense": 0, "income_base": 0, "expense_base": 0, "count": 0}
        if t == "income":
            summary[cur]["income"] = row["total_amount"]
            summary[cur]["income_base"] = row["total_amount_base"]
        elif t == "expense":
            summary[cur]["expense"] = row["total_amount"]
            summary[cur]["expense_base"] = row["total_amount_base"]
        summary[cur]["count"] += row["count"]
        summary_total_count += row["count"]

    # When filtering by a specific account, transfers should count as
    # income (money IN) or expense (money OUT) for that account.
    # Only compute when type filter is absent or explicitly "transfer".
    type_filter = match_query.get("type")
    if account_id and type_filter in (None, "transfer"):
        filtered_acc = await db.accounts.find_one(
            {"id": account_id}, {"_id": 0, "currency": 1}
        )
        acc_currency = filtered_acc.get("currency", "PLN") if filtered_acc else "PLN"
        base_match = {k: v for k, v in match_query.items() if k not in ("$or", "type")}

        if acc_currency not in summary:
            summary[acc_currency] = {"income": 0, "expense": 0, "income_base": 0, "expense_base": 0, "count": 0}

        # Transfer OUT (source account = filtered) → expense
        out_pipeline = [
            {"$match": {**base_match, "account_id": account_id, "type": "transfer"}},
            {"$group": {
                "_id": None,
                "total": {"$sum": {"$ifNull": ["$amount_base", "$amount"]}},
                "count": {"$sum": 1},
            }},
        ]
        out_result = await db.transactions.aggregate(out_pipeline).to_list(1)
        if out_result:
            summary[acc_currency]["expense"] += out_result[0]["total"]
            summary[acc_currency]["expense_base"] += out_result[0]["total"]
            summary[acc_currency]["count"] += out_result[0]["count"]

        # Transfer IN (target account = filtered) → income
        in_pipeline = [
            {"$match": {**base_match, "to_account_id": account_id, "type": "transfer"}},
            {"$group": {
                "_id": None,
                "total": {"$sum": {
                    "$ifNull": ["$to_amount_base", {"$ifNull": ["$amount_base", "$amount"]}]
                }},
                "count": {"$sum": 1},
            }},
        ]
        in_result = await db.transactions.aggregate(in_pipeline).to_list(1)
        if in_result:
            summary[acc_currency]["income"] += in_result[0]["total"]
            summary[acc_currency]["income_base"] += in_result[0]["total"]
            summary[acc_currency]["count"] += in_result[0]["count"]

    # ---- Loans summary (separate block) ----
    # Semantics — ANY operation that changes a loan account balance counts:
    #   Received (Получено) = debt GROWS. Includes:
    #     • transfer where loan acc is SOURCE (took loan → cash/bank)
    #     • direct expense on loan acc (spent borrowed money directly)
    #   Repaid (Погашено)   = debt SHRINKS. Includes:
    #     • transfer where loan acc is TARGET (repaid debt)
    #     • direct income on loan acc (debt partially forgiven / cash returned)
    # NB: pure currency-exchange transfers (is_exchange=True) are excluded —
    # they are NOT loan movements, just conversion of own money.
    loans_summary = None
    if loan_account_ids:
        loan_match_base = {k: v for k, v in match_query.items() if k not in ("$or", "account_id", "to_account_id", "type")}
        loan_match_base["is_exchange"] = {"$ne": True}

        # Received = debt grew (loan acc source for transfer/expense)
        received_pipeline = [
            {"$match": {**loan_match_base,
                        "account_id": {"$in": loan_account_ids},
                        "type": {"$in": ["transfer", "expense"]}}},
            {"$group": {
                "_id": "$account_id",
                "total_base": {"$sum": {"$ifNull": ["$amount_base", "$amount"]}},
                "count": {"$sum": 1},
            }},
        ]
        # Repaid = debt shrunk.
        # 1) transfer where loan acc = TARGET → use to_amount_base
        # 2) direct income on loan acc → use amount_base
        repaid_pipeline = [
            {"$match": {**loan_match_base,
                        "$or": [
                            {"to_account_id": {"$in": loan_account_ids}, "type": "transfer"},
                            {"account_id": {"$in": loan_account_ids}, "type": "income"},
                        ]}},
            {"$addFields": {
                "loan_acc": {"$cond": [
                    {"$eq": ["$type", "income"]}, "$account_id", "$to_account_id"
                ]},
                "effective_base": {"$cond": [
                    {"$eq": ["$type", "income"]},
                    {"$ifNull": ["$amount_base", "$amount"]},
                    {"$ifNull": ["$to_amount_base", {"$ifNull": ["$amount_base", "$amount"]}]},
                ]}
            }},
            {"$group": {
                "_id": "$loan_acc",
                "total_base": {"$sum": "$effective_base"},
                "count": {"$sum": 1},
            }},
        ]
        inflow = await db.transactions.aggregate(received_pipeline).to_list(50)
        outflow = await db.transactions.aggregate(repaid_pipeline).to_list(50)

        # Current loan balance (sum of current_balance for all loan accounts, in PLN-ish)
        loan_accounts_full = await db.accounts.find(
            {"user_id": current_user["user_id"], "is_loan": True, "is_active": {"$ne": False}},
            {"_id": 0, "id": 1, "name": 1, "currency": 1, "current_balance": 1}
        ).to_list(None)

        # Break inflow/outflow down per-currency AND per-account
        # Currency is taken from the LOAN ACCOUNT itself (total_base is already
        # expressed in loan-account currency for both transfers and direct ops).
        received_by_cur: dict = {}
        repaid_by_cur: dict = {}
        per_account: dict = {a["id"]: {
            "id": a["id"], "name": a.get("name"), "currency": a.get("currency"),
            "current_balance": a.get("current_balance", 0),
            "received_by_cur": {}, "received_count": 0,
            "repaid_by_cur": {}, "repaid_count": 0,
        } for a in loan_accounts_full}

        for row in inflow:
            acc = row["_id"]
            pa = per_account.get(acc)
            if not pa:
                continue
            cur = pa.get("currency") or "PLN"
            received_by_cur[cur] = received_by_cur.get(cur, 0) + row["total_base"]
            pa["received_by_cur"][cur] = pa["received_by_cur"].get(cur, 0) + row["total_base"]
            pa["received_count"] += row["count"]

        for row in outflow:
            acc = row["_id"]
            pa = per_account.get(acc)
            if not pa:
                continue
            cur = pa.get("currency") or "PLN"
            repaid_by_cur[cur] = repaid_by_cur.get(cur, 0) + row["total_base"]
            pa["repaid_by_cur"][cur] = pa["repaid_by_cur"].get(cur, 0) + row["total_base"]
            pa["repaid_count"] += row["count"]

        received_total = sum(r["total_base"] for r in inflow)
        repaid_total = sum(r["total_base"] for r in outflow)
        received_count = sum(r["count"] for r in inflow)
        repaid_count = sum(r["count"] for r in outflow)

        loans_summary = {
            "received_base": received_total,
            "received_by_cur": received_by_cur,
            "received_count": received_count,
            "repaid_base": repaid_total,
            "repaid_by_cur": repaid_by_cur,
            "repaid_count": repaid_count,
            "accounts": loan_accounts_full,
            "per_account": list(per_account.values()),
        }
        summary_total_count += loans_summary["received_count"] + loans_summary["repaid_count"]

    # ---- Cash summary (current asset balance, per currency) ----
    # Total money available right NOW across non-loan accounts. Includes loan
    # funds that have already been transferred onto bank/cash accounts.
    asset_accounts = await db.accounts.find(
        {"user_id": current_user["user_id"],
         "is_active": {"$ne": False},
         "is_loan": {"$ne": True}},
        {"_id": 0, "id": 1, "name": 1, "currency": 1, "current_balance": 1,
         "initial_balance": 1}
    ).to_list(None)
    cash_by_cur: dict = {}
    for a in asset_accounts:
        cur = a.get("currency") or "PLN"
        cash_by_cur[cur] = cash_by_cur.get(cur, 0) + (a.get("current_balance") or 0)

    # Period-start and period-end balances (only if date filter is set).
    # Computed by replaying transactions onto initial_balance up to the chosen
    # cut-off date for each account.
    period_start_by_cur: dict = {}
    period_end_by_cur: dict = {}
    period_start_date = None
    period_end_date = None
    if date_from or date_to:
        from datetime import date as _date
        period_start_date = date_from
        period_end_date = date_to or _date.today().isoformat()

        # Compute "balance at end of date X" for each asset account
        async def _balances_at(cutoff_date: str) -> dict:
            balances = {a["id"]: float(a.get("initial_balance") or 0) for a in asset_accounts}
            asset_ids = list(balances.keys())
            txs = await db.transactions.find(
                {"user_id": current_user["user_id"],
                 "status": "fact",
                 "date": {"$lte": cutoff_date},
                 "$or": [
                     {"account_id": {"$in": asset_ids}},
                     {"to_account_id": {"$in": asset_ids}},
                 ]},
                {"_id": 0, "type": 1, "account_id": 1, "to_account_id": 1,
                 "amount_base": 1, "to_amount_base": 1, "amount": 1}
            ).to_list(50000)
            for t in txs:
                acc = t.get("account_id")
                to_acc = t.get("to_account_id")
                amt = t.get("amount_base") if t.get("amount_base") is not None else t.get("amount", 0)
                ttype = t.get("type")
                if ttype == "income" and acc in balances:
                    balances[acc] += amt
                elif ttype == "expense" and acc in balances:
                    balances[acc] -= amt
                elif ttype == "transfer":
                    if acc in balances:
                        balances[acc] -= amt
                    if to_acc in balances:
                        to_amt = t.get("to_amount_base") if t.get("to_amount_base") is not None else amt
                        balances[to_acc] += to_amt
            return balances

        if period_start_date:
            # Balance at the END of (start_date - 1 day) = balance ENTERING the period
            from datetime import datetime as _dt
            try:
                _sd = _dt.fromisoformat(period_start_date).date()
                prev_day = (_sd - timedelta(days=1)).isoformat()
            except Exception:
                prev_day = period_start_date
            start_bal = await _balances_at(prev_day)
            for a in asset_accounts:
                cur = a.get("currency") or "PLN"
                period_start_by_cur[cur] = period_start_by_cur.get(cur, 0) + start_bal.get(a["id"], 0)

        end_bal = await _balances_at(period_end_date)
        for a in asset_accounts:
            cur = a.get("currency") or "PLN"
            period_end_by_cur[cur] = period_end_by_cur.get(cur, 0) + end_bal.get(a["id"], 0)

    cash_summary = {
        "by_currency": cash_by_cur,
        "period_start_by_currency": period_start_by_cur or None,
        "period_end_by_currency": period_end_by_cur or None,
        "period_start_date": period_start_date,
        "period_end_date": period_end_date,
        "accounts": [
            {k: v for k, v in a.items() if k != "initial_balance"}
            for a in asset_accounts
        ],
    }

    return {
        "items": items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page if per_page else 1,
        "summary": summary,
        "summary_total_count": summary_total_count,
        "loans_summary": loans_summary,
        "cash_summary": cash_summary,
    }


@router.post("/transactions", response_model=Transaction)
async def create_transaction(data: TransactionCreate, current_user: dict = Depends(get_current_user)):
    category_name = None
    if data.category_id:
        cat = await db.categories.find_one({"id": data.category_id}, {"_id": 0, "name": 1})
        category_name = cat["name"] if cat else None

    direction = await db.directions.find_one({"id": data.direction_id}, {"_id": 0, "name": 1})
    direction_name = direction["name"] if direction else None

    account = await db.accounts.find_one({"id": data.account_id}, {"_id": 0, "name": 1, "current_balance": 1, "currency": 1})
    account_name = account["name"] if account else None

    # Calculate converted amount if currencies differ
    amount_base, exchange_rate = await calc_amount_base(
        data.amount, data.currency, data.account_id, current_user["user_id"]
    )

    # For transfers: calculate amount in target account's currency
    to_amount_base = None
    to_account_name = None
    if data.to_account_id:
        to_acc = await db.accounts.find_one({"id": data.to_account_id}, {"_id": 0, "name": 1, "currency": 1})
        to_account_name = to_acc["name"] if to_acc else None
        to_acc_currency = (to_acc or {}).get("currency", "PLN")
        # If user provided a manual to_amount AND currencies differ → use it as override
        if data.to_amount is not None and data.to_amount > 0 and to_acc_currency != data.currency:
            to_amount_base = float(data.to_amount)
            if data.amount and data.amount != 0:
                # Convention: rate = to_amount / amount (target units per 1 source unit)
                exchange_rate = round(data.to_amount / data.amount, 6)
        else:
            to_amount_base_val, _ = await calc_amount_base(
                data.amount, data.currency, data.to_account_id, current_user["user_id"]
            )
            to_amount_base = to_amount_base_val

    contractor_name = None
    if data.contractor_id:
        contractor = await db.contractors.find_one({"id": data.contractor_id}, {"_id": 0, "name": 1})
        contractor_name = contractor["name"] if contractor else None

    transaction = Transaction(
        **data.model_dump(),
        user_id=current_user["user_id"],
        category_name=category_name,
        direction_name=direction_name,
        account_name=account_name,
        to_account_name=to_account_name,
        contractor_name=contractor_name,
        amount_base=amount_base,
        to_amount_base=to_amount_base,
        exchange_rate=exchange_rate,
        source="manual"
    )

    await db.transactions.insert_one(transaction.model_dump())

    await update_account_balance(data.account_id, current_user["user_id"])
    if data.to_account_id:
        await update_account_balance(data.to_account_id, current_user["user_id"])

    updated_account = await db.accounts.find_one({"id": data.account_id}, {"_id": 0, "current_balance": 1})
    transaction.balance_after = updated_account["current_balance"] if updated_account else 0

    await db.transactions.update_one(
        {"id": transaction.id},
        {"$set": {"balance_after": transaction.balance_after}}
    )

    return transaction


@router.put("/transactions/{transaction_id}", response_model=Transaction)
async def update_transaction(transaction_id: str, data: TransactionCreate, current_user: dict = Depends(get_current_user)):
    old_transaction = await db.transactions.find_one({"id": transaction_id, "user_id": current_user["user_id"]}, {"_id": 0})
    if not old_transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")

    category_name = None
    if data.category_id:
        cat = await db.categories.find_one({"id": data.category_id}, {"_id": 0, "name": 1})
        category_name = cat["name"] if cat else None

    direction = await db.directions.find_one({"id": data.direction_id}, {"_id": 0, "name": 1})
    direction_name = direction["name"] if direction else None

    account = await db.accounts.find_one({"id": data.account_id}, {"_id": 0, "name": 1})
    account_name = account["name"] if account else None

    # Calculate converted amount
    amount_base, exchange_rate = await calc_amount_base(
        data.amount, data.currency, data.account_id, current_user["user_id"]
    )

    to_amount_base = None
    to_account_name = None
    if data.to_account_id:
        to_acc = await db.accounts.find_one({"id": data.to_account_id}, {"_id": 0, "name": 1, "currency": 1})
        to_account_name = to_acc["name"] if to_acc else None
        to_acc_currency = (to_acc or {}).get("currency", "PLN")
        if data.to_amount is not None and data.to_amount > 0 and to_acc_currency != data.currency:
            to_amount_base = float(data.to_amount)
            if data.amount and data.amount != 0:
                exchange_rate = round(data.to_amount / data.amount, 6)
        else:
            to_amount_base_val, _ = await calc_amount_base(
                data.amount, data.currency, data.to_account_id, current_user["user_id"]
            )
            to_amount_base = to_amount_base_val

    contractor_name = None
    if data.contractor_id:
        contractor = await db.contractors.find_one({"id": data.contractor_id}, {"_id": 0, "name": 1})
        contractor_name = contractor["name"] if contractor else None

    update_data = data.model_dump()
    update_data["category_name"] = category_name
    update_data["direction_name"] = direction_name
    update_data["account_name"] = account_name
    update_data["to_account_name"] = to_account_name
    update_data["contractor_name"] = contractor_name
    update_data["amount_base"] = amount_base
    update_data["to_amount_base"] = to_amount_base
    update_data["exchange_rate"] = exchange_rate

    await db.transactions.update_one(
        {"id": transaction_id, "user_id": current_user["user_id"]},
        {"$set": update_data}
    )

    await update_account_balance(data.account_id, current_user["user_id"])
    if old_transaction.get("account_id") != data.account_id:
        await update_account_balance(old_transaction["account_id"], current_user["user_id"])
    if data.to_account_id:
        await update_account_balance(data.to_account_id, current_user["user_id"])
    # If the OLD record had a to_account (was a transfer) and the new one doesn't,
    # or the new to_account differs — refresh the old to_account too so balance
    # reflects the change of type/destination.
    old_to = old_transaction.get("to_account_id")
    if old_to and old_to != data.to_account_id:
        await update_account_balance(old_to, current_user["user_id"])

    transaction = await db.transactions.find_one({"id": transaction_id}, {"_id": 0})
    return transaction


@router.delete("/transactions/{transaction_id}")
async def delete_transaction(transaction_id: str, current_user: dict = Depends(get_current_user)):
    transaction = await db.transactions.find_one({"id": transaction_id, "user_id": current_user["user_id"]}, {"_id": 0})
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")

    await db.transactions.delete_one({"id": transaction_id})

    await update_account_balance(transaction["account_id"], current_user["user_id"])
    if transaction.get("to_account_id"):
        await update_account_balance(transaction["to_account_id"], current_user["user_id"])

    return {"status": "deleted"}


@router.post("/transactions/bulk-delete")
async def bulk_delete_transactions(
    payload: dict,
    current_user: dict = Depends(get_current_user),
):
    """Delete multiple transactions at once. Body: {"ids": ["id1", "id2", ...]}."""
    ids = payload.get("ids") or []
    if not isinstance(ids, list) or not ids:
        raise HTTPException(status_code=400, detail="Передайте массив ids")
    if len(ids) > 500:
        raise HTTPException(status_code=400, detail="За один раз можно удалить до 500 операций")

    # Collect affected accounts so we can refresh balances after deletion
    to_refresh: set[str] = set()
    cursor = db.transactions.find(
        {"id": {"$in": ids}, "user_id": current_user["user_id"]},
        {"_id": 0, "id": 1, "account_id": 1, "to_account_id": 1}
    )
    matched_ids = []
    async for t in cursor:
        matched_ids.append(t["id"])
        if t.get("account_id"):
            to_refresh.add(t["account_id"])
        if t.get("to_account_id"):
            to_refresh.add(t["to_account_id"])

    if not matched_ids:
        return {"status": "ok", "deleted": 0}

    res = await db.transactions.delete_many({
        "id": {"$in": matched_ids},
        "user_id": current_user["user_id"]
    })

    for acc_id in to_refresh:
        await update_account_balance(acc_id, current_user["user_id"])

    return {"status": "ok", "deleted": res.deleted_count}


@router.post("/transactions/bulk-update")
async def bulk_update_transactions(
    payload: dict,
    current_user: dict = Depends(get_current_user),
):
    """Bulk update category and/or direction for multiple transactions.
    Body: {"ids": [...], "category_id": "...", "direction_id": "..."}
    At least one of category_id / direction_id must be provided.
    """
    ids = payload.get("ids") or []
    category_id = payload.get("category_id")
    direction_id = payload.get("direction_id")

    if not isinstance(ids, list) or not ids:
        raise HTTPException(status_code=400, detail="Передайте массив ids")
    if len(ids) > 500:
        raise HTTPException(status_code=400, detail="За один раз можно обновить до 500 операций")
    if not category_id and not direction_id:
        raise HTTPException(status_code=400, detail="Укажите category_id и/или direction_id")

    update_fields: dict = {}

    if category_id:
        cat = await db.categories.find_one(
            {"id": category_id, "user_id": current_user["user_id"]},
            {"_id": 0, "name": 1}
        )
        if not cat:
            raise HTTPException(status_code=404, detail="Статья не найдена")
        update_fields["category_id"] = category_id
        update_fields["category_name"] = cat["name"]

    if direction_id:
        d = await db.directions.find_one(
            {"id": direction_id, "user_id": current_user["user_id"]},
            {"_id": 0, "name": 1}
        )
        if not d:
            raise HTTPException(status_code=404, detail="Направление не найдено")
        update_fields["direction_id"] = direction_id

    res = await db.transactions.update_many(
        {"id": {"$in": ids}, "user_id": current_user["user_id"]},
        {"$set": update_fields}
    )

    return {"status": "ok", "matched": res.matched_count, "modified": res.modified_count}


@router.post("/transactions/bulk-apply-rules")
async def bulk_apply_auto_rules(
    payload: dict,
    current_user: dict = Depends(get_current_user),
):
    """Apply auto-categorization rules to selected transactions.
    Body: {"ids": [...], "overwrite": false}
    - When overwrite=False (default), only fills missing category_id/direction_id.
    - When overwrite=True, overrides existing values where a matching rule fires.
    """
    ids = payload.get("ids") or []
    overwrite = bool(payload.get("overwrite", False))
    if not isinstance(ids, list) or not ids:
        raise HTTPException(status_code=400, detail="Передайте массив ids")
    if len(ids) > 500:
        raise HTTPException(status_code=400, detail="За один раз можно обработать до 500 операций")

    rules = await db.auto_rules.find(
        {"user_id": current_user["user_id"], "is_active": True},
        {"_id": 0}
    ).to_list(500)
    if not rules:
        return {"status": "ok", "matched": 0, "updated": 0, "skipped": len(ids), "no_match": len(ids)}

    # Preload categories/directions for fast lookup
    cat_ids = {r.get("category_id") for r in rules if r.get("category_id")}
    dir_ids = {r.get("direction_id") for r in rules if r.get("direction_id")}
    cats = {c["id"]: c["name"] for c in await db.categories.find(
        {"id": {"$in": list(cat_ids)}}, {"_id": 0, "id": 1, "name": 1}
    ).to_list(None)} if cat_ids else {}
    dirs = {d["id"]: d["name"] for d in await db.directions.find(
        {"id": {"$in": list(dir_ids)}}, {"_id": 0, "id": 1, "name": 1}
    ).to_list(None)} if dir_ids else {}

    txs = await db.transactions.find(
        {"id": {"$in": ids}, "user_id": current_user["user_id"]},
        {"_id": 0, "id": 1, "description": 1, "category_id": 1, "direction_id": 1}
    ).to_list(len(ids))

    updated = 0
    no_match = 0
    skipped = 0
    for t in txs:
        desc = (t.get("description") or "").lower()
        match = None
        for rule in rules:
            pat = (rule.get("pattern") or "").lower().strip()
            if pat and pat in desc:
                match = rule
                break
        if not match:
            no_match += 1
            continue

        updates = {}
        new_cat = match.get("category_id")
        new_dir = match.get("direction_id")

        if new_cat and (overwrite or not t.get("category_id")):
            updates["category_id"] = new_cat
            updates["category_name"] = cats.get(new_cat)
        if new_dir and (overwrite or not t.get("direction_id")):
            updates["direction_id"] = new_dir
            updates["direction_name"] = dirs.get(new_dir)

        if updates:
            await db.transactions.update_one({"id": t["id"]}, {"$set": updates})
            updated += 1
        else:
            skipped += 1

    return {
        "status": "ok",
        "matched": updated + skipped,
        "updated": updated,
        "skipped": skipped,
        "no_match": no_match,
    }


@router.post("/import/preview")
async def preview_import(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    content = await file.read()

    rows = []
    if file.filename.endswith('.csv'):
        text = content.decode('utf-8')
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    elif file.filename.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                rows.append(dict(zip(headers, row)))

    columns = list(rows[0].keys()) if rows else []

    return {
        "columns": columns,
        "preview": rows[:100],
        "total_rows": len(rows)
    }


@router.post("/import/process")
async def process_import(
    file: UploadFile = File(...),
    date_column: str = Query(...),
    amount_column: str = Query(...),
    description_column: str = Query(...),
    type_column: Optional[str] = None,
    account_id: str = Query(...),
    direction_id: str = Query(...),
    current_user: dict = Depends(get_current_user)
):
    content = await file.read()

    rows = []
    if file.filename.endswith('.csv'):
        text = content.decode('utf-8')
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    elif file.filename.endswith('.xlsx'):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                rows.append(dict(zip(headers, row)))

    rules = await db.auto_rules.find({"user_id": current_user["user_id"], "is_active": True}, {"_id": 0}).to_list(100)

    account = await db.accounts.find_one({"id": account_id}, {"_id": 0, "name": 1})
    direction = await db.directions.find_one({"id": direction_id}, {"_id": 0, "name": 1})

    imported = []
    duplicates = []

    for row in rows:
        try:
            date_val = str(row.get(date_column, ""))
            amount_val = row.get(amount_column, 0)
            description = str(row.get(description_column, ""))

            if isinstance(date_val, str) and date_val:
                for fmt in ["%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d"]:
                    try:
                        parsed_date = datetime.strptime(date_val.strip(), fmt)
                        date_str = parsed_date.strftime("%Y-%m-%d")
                        break
                    except ValueError:
                        date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            else:
                date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")

            if isinstance(amount_val, str):
                amount_val = amount_val.replace(",", ".").replace(" ", "")
            amount = abs(float(amount_val))

            trans_type = "expense"
            if type_column and row.get(type_column):
                type_val = str(row.get(type_column, "")).lower()
                if "income" in type_val or "приход" in type_val or "+" in type_val:
                    trans_type = "income"
            elif float(amount_val) > 0:
                trans_type = "income"

            existing = await db.transactions.find_one({
                "user_id": current_user["user_id"],
                "date": date_str,
                "amount": amount,
                "description": description
            }, {"_id": 0})

            if existing:
                duplicates.append({"date": date_str, "amount": amount, "description": description})
                continue

            category_id = None
            category_name = None
            matched_direction_id = direction_id
            matched_direction_name = direction["name"] if direction else None
            matched = False

            for rule in rules:
                if rule["pattern"].lower() in description.lower():
                    if rule.get("category_id"):
                        category_id = rule["category_id"]
                        cat = await db.categories.find_one({"id": category_id}, {"_id": 0, "name": 1})
                        category_name = cat["name"] if cat else None
                    if rule.get("direction_id"):
                        matched_direction_id = rule["direction_id"]
                        dir_doc = await db.directions.find_one({"id": matched_direction_id}, {"_id": 0, "name": 1})
                        matched_direction_name = dir_doc["name"] if dir_doc else None
                    matched = True
                    break

            transaction = Transaction(
                date=date_str,
                type=trans_type,
                amount=amount,
                currency="PLN",
                category_id=category_id,
                category_name=category_name,
                direction_id=matched_direction_id,
                direction_name=matched_direction_name,
                account_id=account_id,
                account_name=account["name"] if account else None,
                description=description,
                source="import",
                status="fact",
                user_id=current_user["user_id"]
            )

            await db.transactions.insert_one(transaction.model_dump())
            imported.append({
                "id": transaction.id,
                "date": date_str,
                "type": trans_type,
                "amount": amount,
                "description": description,
                "category_name": category_name,
                "direction_name": matched_direction_name,
                "matched": matched
            })

        except Exception as e:
            logger.error(f"Error importing row: {e}")
            continue

    await update_account_balance(account_id, current_user["user_id"])

    return {
        "imported_count": len(imported),
        "duplicate_count": len(duplicates),
        "imported": imported,
        "duplicates": duplicates
    }



@router.get("/transactions/descriptions/suggestions")
async def get_description_suggestions(
    q: str = "",
    current_user: dict = Depends(get_current_user)
):
    """Return popular transaction descriptions for autocomplete."""
    query = {"user_id": current_user["user_id"], "description": {"$exists": True, "$ne": ""}}
    if q:
        query["description"] = {"$regex": q, "$options": "i"}

    pipeline = [
        {"$match": query},
        {"$group": {"_id": "$description", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 20},
    ]
    results = await db.transactions.aggregate(pipeline).to_list(20)
    return [{"description": r["_id"], "count": r["count"]} for r in results if r["_id"]]


@router.put("/transactions/{transaction_id}/review")
async def toggle_needs_review(
    transaction_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Toggle the needs_review flag on a transaction."""
    tx = await db.transactions.find_one(
        {"id": transaction_id, "user_id": current_user["user_id"]},
        {"_id": 0, "needs_review": 1}
    )
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")

    new_val = not tx.get("needs_review", False)
    await db.transactions.update_one(
        {"id": transaction_id},
        {"$set": {"needs_review": new_val}}
    )
    return {"needs_review": new_val}
